"""Part Estimation — cost the spare parts a lot needs before production starts.

Lots Summary lists every lot with its tag count and how many Required=Yes part
rows its IQC data produces. Generate Estimate opens a costing modal (quantity
per part comes from IQC, unit cost is entered by the user); Download & Attach
Estimate writes an Excel workbook, attaches it to the lot row, and raises a
PartSourcingRequest so the request lands on Part Master -> Sourcing Requests.
"""
import io
import os
import uuid
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse
from sqlalchemy import select, desc
from sqlalchemy.ext.asyncio import AsyncSession

from database import get_db
from templates_config import templates
from utils.timezone import app_now
from models.user import User
from models.lot import Lot
from models.part_estimate import PartEstimate, PartEstimateLine
from models.part_request import PartSourcingRequest
from auth.dependencies import require_module_perm, verify_csrf
from services.audit_engine import audit
from services.part_estimation import (
    device_counts_by_lot, required_counts_by_lot, required_parts_for_lot,
)
from services.part_estimate_matrix import (
    PART_GROUPS, group_meta, lot_model_matrix, quantities_for,
)
from utils.grades import grade_options

router = APIRouter(prefix="/part-estimation", tags=["part_estimation"],
                   dependencies=[Depends(verify_csrf)])
allowed = require_module_perm("part_estimation")

UPLOAD_DIR = "uploads/part_estimates"


def _as_uuid(val):
    try:
        return uuid.UUID(str(val))
    except (ValueError, AttributeError, TypeError):
        raise HTTPException(400, "Invalid id")


def _money(val) -> Decimal:
    try:
        d = Decimal(str(val or 0))
    except (InvalidOperation, ValueError):
        return Decimal("0")
    return d.quantize(Decimal("0.01"))


@router.get("", response_class=HTMLResponse)
async def part_estimation_page(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(allowed),
):
    lots = (await db.execute(
        select(Lot).where(Lot.is_trashed == False).order_by(desc(Lot.purchase_date))  # noqa: E712
    )).scalars().all()

    tag_counts = await device_counts_by_lot(db)
    parts_counts = await required_counts_by_lot(db)

    # Every attached estimate per lot, newest first — a lot can be estimated
    # more than once (re-priced, or one Generate click across several grades),
    # and the Download column now offers all of them, not just the latest.
    estimates = (await db.execute(
        select(PartEstimate)
        .where(PartEstimate.file_name.isnot(None))
        .order_by(desc(PartEstimate.created_at), desc(PartEstimate.id))
    )).scalars().all()
    estimates_by_lot = {}
    for e in estimates:
        estimates_by_lot.setdefault(e.lot_id, []).append(e)

    rows = [{
        "lot": lot,
        "total_products": tag_counts.get(lot.id, 0),
        "total_parts_required": parts_counts.get(lot.id, 0),
        "estimates": estimates_by_lot.get(lot.id, []),
    } for lot in lots]

    return templates.TemplateResponse("part_estimation/index.html", {
        "request": request, "current_user": current_user, "rows": rows,
        "grade_options": grade_options(),
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error"),
    })


@router.get("/{lot_id}/parts")
async def lot_parts(
    lot_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(allowed),
):
    """Modal payload: one row per required part, plus the unit costs from this
    lot's previous estimate so a re-estimate starts from what was priced last."""
    lid = _as_uuid(lot_id)
    lot = (await db.execute(select(Lot).where(Lot.id == lid))).scalar_one_or_none()
    if not lot:
        raise HTTPException(404, "Lot not found")

    # The modal lists every Parts Consumption row, including the ones no tag in
    # this lot needs, so the operator can see the full checklist rather than
    # wonder whether a missing part was overlooked. Zero rows carry a 0 and are
    # dropped when the estimate is saved.
    parts = await required_parts_for_lot(db, lid, include_zero=True)

    prev = (await db.execute(
        select(PartEstimate).where(PartEstimate.lot_id == lid)
        .order_by(desc(PartEstimate.created_at)).limit(1)
    )).scalar_one_or_none()
    prev_costs, prev_labour, prev_notes = {}, "0", ""
    if prev:
        prev_lines = (await db.execute(
            select(PartEstimateLine).where(PartEstimateLine.estimate_id == prev.id)
        )).scalars().all()
        prev_costs = {l.part_name: str(l.unit_cost) for l in prev_lines}
        prev_labour = str(prev.labour_cost or 0)
        prev_notes = prev.notes or ""

    for p in parts:
        p["unit_cost"] = prev_costs.get(p["part_name"], "")

    return JSONResponse({
        "ok": True, "lot_number": lot.lot_number, "parts": parts,
        "labour_cost": prev_labour, "notes": prev_notes,
    })


@router.get("/{lot_id}/model-matrix")
async def lot_model_matrix_json(
    lot_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(allowed),
):
    """Payload for the Part Estimate modal: the lot's Model Summary rows, each
    with a status breakdown per part, plus the last prices entered for this lot
    so a re-estimate starts where the previous one finished."""
    lid = _as_uuid(lot_id)
    lot = (await db.execute(select(Lot).where(Lot.id == lid))).scalar_one_or_none()
    if not lot:
        raise HTTPException(404, "Lot not found")

    models = await lot_model_matrix(db, lid)

    prev = (await db.execute(
        select(PartEstimate).where(PartEstimate.lot_id == lid)
        .order_by(desc(PartEstimate.created_at), desc(PartEstimate.id)).limit(1)
    )).scalar_one_or_none()
    prev_costs, prev_labour, prev_notes = {}, "0", ""
    if prev:
        prev_lines = (await db.execute(
            select(PartEstimateLine).where(PartEstimateLine.estimate_id == prev.id)
        )).scalars().all()
        # Lines from the per-model estimate are stored as "<model> — <part>";
        # keying on the whole label lets a re-estimate re-fill the exact cell,
        # and lines from a plain Generate Estimate simply never match. When a
        # batch wrote several grade files, they are byte-identical in content
        # (see generate_model_estimate), so it does not matter which one of
        # them "prev" resolves to.
        prev_costs = {l.part_name: str(l.unit_cost) for l in prev_lines}
        prev_labour = str(prev.labour_cost or 0)
        prev_notes = prev.notes or ""

    return JSONResponse({
        "ok": True, "lot_number": lot.lot_number, "groups": group_meta(),
        "models": models, "prev_costs": prev_costs,
        "labour_cost": prev_labour, "notes": prev_notes,
        "grade_options": grade_options(),
    })


def _line_label(model: str, part: str) -> str:
    """PartEstimateLine.part_name is 150 chars; a long model name must not
    cost us the part name at the end of it."""
    label = f"{model} — {part}"
    if len(label) <= 150:
        return label
    return f"{model[:150 - len(part) - 6]}… — {part}"


def _build_model_workbook(*, lot_number, generated_on, generated_by, grade, blocks,
                          unit_price_total, parts_total, labour_cost,
                          total_estimation, notes) -> bytes:
    """Workbook for the per-model estimate: Sheet 1 is the detailed breakdown
    (one block per model, one section per group — what the operator priced on
    screen); Sheet 2 is a compact roll-up, one row per model. Rows with
    Parts Quantity 0 are left out of both sheets — a status band that needs
    nothing costs nothing and would only pad the workbook.

    `grade` is stamped in the header of both sheets: when the operator picks
    more than one grade, every other field in this workbook is identical
    across the batch — grade is the one thing that tells the files apart.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Part Estimate"

    head_fill = PatternFill("solid", fgColor="1F3864")
    model_fill = PatternFill("solid", fgColor="DCE6F1")
    group_fill = PatternFill("solid", fgColor="F2F2F2")
    head_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money = '#,##0.00'

    def _header(sheet, title):
        sheet["A1"] = title
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A2"] = f"Lot Number: {lot_number}"
        sheet["A2"].font = bold
        sheet["A3"] = f"Grade: {grade or '—'}"
        sheet["A3"].font = bold
        sheet["A4"] = f"Generated On: {generated_on}"
        sheet["A5"] = f"Generated By: {generated_by}"

    _header(ws, "Part Estimate")

    headers = ["Part", "Note", "Parts Quantity", "Unit Part Price", "Total Part Value"]
    r = 7
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=r, column=col, value=title)
        c.fill, c.font, c.border = head_fill, head_font, border
        c.alignment = Alignment(horizontal="center")

    for block in blocks:
        # A model whose every group is costing 0 (nothing ticked, or every
        # group left on "Select") contributes nothing to either sheet.
        if not block["qty"]:
            continue
        r += 1
        c = ws.cell(row=r, column=1,
                    value=f"{block['model']} ({block['make']} · {block['device_type']})")
        c.font = bold
        ws.cell(row=r, column=2, value=f"Model Count: {block['model_count']}")
        ws.cell(row=r, column=3, value=block["qty"]).font = bold
        ws.cell(row=r, column=4, value=float(block["unit_price_sum"])).number_format = money
        cv = ws.cell(row=r, column=5, value=float(block["total"]))
        cv.font, cv.number_format = bold, money
        for col in range(1, 6):
            ws.cell(row=r, column=col).fill = model_fill
            ws.cell(row=r, column=col).border = border

        for grp in block["groups"]:
            lines = [l for l in grp["lines"] if l["qty"]]
            if not lines:
                continue
            r += 1
            g = ws.cell(row=r, column=1, value=f"    {grp['heading']}")
            g.font = bold
            ws.cell(row=r, column=2, value=f"{grp['filter_label']}: {grp['status'] or 'Select'}")
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = group_fill
                ws.cell(row=r, column=col).border = border
            for line in lines:
                r += 1
                ws.cell(row=r, column=1, value=f"        {line['part']}").border = border
                ws.cell(row=r, column=2, value=line["note"] or "").border = border
                ws.cell(row=r, column=3, value=line["qty"]).border = border
                c4 = ws.cell(row=r, column=4, value=float(line["unit_cost"]))
                c4.border, c4.number_format = border, money
                c5 = ws.cell(row=r, column=5, value=float(line["total_cost"]))
                c5.border, c5.number_format = border, money

    r += 2
    for label, value, fmt in (
        ("Total Unit Price", float(unit_price_total), money),
        ("Total Parts Price", float(parts_total), money),
        ("Total Labour Cost", float(labour_cost), money),
        ("Total Estimate Cost", float(total_estimation), money),
    ):
        ws.cell(row=r, column=1, value=label).font = bold
        c = ws.cell(row=r, column=5, value=value)
        c.font, c.number_format = bold, fmt
        r += 1

    if notes:
        r += 1
        ws.cell(row=r, column=1, value="Notes").font = bold
        nc = ws.cell(row=r, column=2, value=notes)
        nc.alignment = Alignment(wrap_text=True, vertical="top")

    for col, width in (("A", 40), ("B", 34), ("C", 16), ("D", 18), ("E", 20)):
        ws.column_dimensions[col].width = width

    # ── Sheet 2 — Model Summary: one row per priced model ─────────────────────
    ms = wb.create_sheet("Model Summary")
    _header(ms, "Model Summary")
    ms_headers = ["Model", "Make", "Device Type", "Model Count",
                  "Parts Quantity", "Unit Price", "Total Parts Price"]
    r2 = 7
    for col, title in enumerate(ms_headers, start=1):
        c = ms.cell(row=r2, column=col, value=title)
        c.fill, c.font, c.border = head_fill, head_font, border
        c.alignment = Alignment(horizontal="center")
    total_model_qty = 0
    for block in blocks:
        if not block["qty"]:
            continue
        r2 += 1
        total_model_qty += block["qty"]
        ms.cell(row=r2, column=1, value=block["model"]).border = border
        ms.cell(row=r2, column=2, value=block["make"]).border = border
        ms.cell(row=r2, column=3, value=block["device_type"]).border = border
        ms.cell(row=r2, column=4, value=block["model_count"]).border = border
        ms.cell(row=r2, column=5, value=block["qty"]).border = border
        c6 = ms.cell(row=r2, column=6, value=float(block["unit_price_sum"]))
        c6.border, c6.number_format = border, money
        c7 = ms.cell(row=r2, column=7, value=float(block["total"]))
        c7.border, c7.number_format = border, money
    r2 += 1
    ms.cell(row=r2, column=1, value="Total").font = bold
    ms.cell(row=r2, column=5, value=total_model_qty).font = bold
    tc = ms.cell(row=r2, column=7, value=float(parts_total))
    tc.font, tc.number_format = bold, money
    for col in range(1, 8):
        ms.cell(row=r2, column=col).border = border
    for col, width in (("A", 30), ("B", 18), ("C", 16), ("D", 14), ("E", 16), ("F", 16), ("G", 18)):
        ms.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.post("/{lot_id}/generate-model-estimate")
async def generate_model_estimate(
    lot_id: str,
    request: Request,
    payload: str = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(allowed),
):
    """Part Estimate -> Generate Estimate.

    payload = {
      included: [model_key, ...],
      statuses: {model_key: {group: [status, ...]}},  # multi-select per group
      prices:   {model_key: {"<group>|<part>": {note, unit_cost}}},
      grade:    grade_value,
      labour_cost, notes
    }

    Quantities are recomputed here from IQC; the browser only gets to choose
    the status bands, the notes and the prices. Same rule as the plain
    Generate Estimate — a tampered or stale quantity can never reach the
    workbook or the sourcing request.

    One call writes one workbook for one grade. The modal stays open after a
    successful call specifically so the operator can change the grade (and
    re-price whatever that grade needs differently) and call this again —
    each call is its own independent PartEstimate + its own
    PartSourcingRequest, exactly like the plain (non-model) Generate Estimate
    already works. A lot ends up with as many files as the operator chose to
    make, in as many separate visits to this endpoint as that took.
    """
    import json

    lid = _as_uuid(lot_id)
    lot = (await db.execute(select(Lot).where(Lot.id == lid))).scalar_one_or_none()
    if not lot:
        raise HTTPException(404, "Lot not found")

    try:
        data = json.loads(payload)
    except (ValueError, TypeError):
        return JSONResponse({"ok": False, "error": "Malformed estimate payload"}, status_code=400)

    matrix = await lot_model_matrix(db, lid)
    if not matrix:
        return JSONResponse({"ok": False, "error": "This lot has no tags to estimate"},
                            status_code=400)

    included = set(data.get("included") or [])
    statuses = data.get("statuses") or {}
    prices = data.get("prices") or {}
    grade = (data.get("grade") or "").strip()
    if not grade:
        return JSONResponse({"ok": False, "error": "Select a Grade"}, status_code=400)
    grade_labels = dict(grade_options())
    if grade not in grade_labels:
        return JSONResponse({"ok": False, "error": f"Unknown grade: {grade}"}, status_code=400)
    grade_label = grade_labels[grade]

    qtys = quantities_for(matrix, statuses)

    blocks, lines = [], []
    unit_price_total = Decimal("0")
    parts_total = Decimal("0")
    total_qty = 0

    for row in matrix:
        if row["key"] not in included:
            continue
        model_prices = prices.get(row["key"]) or {}
        block = {"model": row["model"], "make": row["make"],
                 "device_type": row["device_type"], "model_count": row["qty"],
                 "qty": 0, "unit_price_sum": Decimal("0"), "total": Decimal("0"),
                 "groups": []}

        for g, heading, filter_label, group_statuses, parts in PART_GROUPS:
            picked = [s for s in ((statuses.get(row["key"]) or {}).get(g) or [])
                     if s in group_statuses]
            grp = {"heading": heading, "filter_label": filter_label,
                   "status": ", ".join(picked) or None, "lines": []}
            for name, _fn in parts:
                cell = model_prices.get(f"{g}|{name}") or {}
                qty = qtys[row["key"]][g][name]
                unit = _money(cell.get("unit_cost", 0))
                total = _money(unit * qty)
                note = (str(cell.get("note") or "")).strip()[:255] or None
                grp["lines"].append({"part": name, "note": note, "qty": qty,
                                     "unit_cost": unit, "total_cost": total})
                block["qty"] += qty
                block["unit_price_sum"] += unit
                block["total"] += total
                unit_price_total += unit
                parts_total += total
                total_qty += qty
                if qty:
                    lines.append({"part_name": _line_label(row["model"], name),
                                  "qty": qty, "unit_cost": unit, "total_cost": total})
            block["groups"].append(grp)
        blocks.append(block)

    priced_blocks = [b for b in blocks if b["qty"]]
    if not blocks:
        return JSONResponse({"ok": False, "error": "Select at least one model to estimate"},
                            status_code=400)
    if not priced_blocks:
        return JSONResponse(
            {"ok": False, "error": "Every selected group is costing 0 parts — "
                                   "choose a part status before generating"},
            status_code=400)

    labour = _money(data.get("labour_cost"))
    total_estimation = _money(parts_total + labour)
    notes = (data.get("notes") or "").strip() or None

    total_products = (await device_counts_by_lot(db)).get(lid, 0)
    stamp = app_now()
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    safe_lot = "".join(ch for ch in (lot.lot_number or "lot") if ch.isalnum() or ch in "-_")

    xlsx = _build_model_workbook(
        lot_number=lot.lot_number, generated_on=stamp.strftime("%d %b %Y %H:%M"),
        generated_by=current_user.username, grade=grade_label, blocks=blocks,
        unit_price_total=unit_price_total, parts_total=parts_total,
        labour_cost=labour, total_estimation=total_estimation, notes=notes,
    )
    safe_grade = "".join(ch for ch in grade_label if ch.isalnum() or ch in "-_") or "grade"
    stored = f"{uuid.uuid4().hex}.xlsx"
    download_name = f"Part_Estimate_{safe_lot}_{safe_grade}_{stamp.strftime('%Y%m%d-%H%M')}.xlsx"
    with open(os.path.join(UPLOAD_DIR, stored), "wb") as fh:
        fh.write(xlsx)

    est = PartEstimate(
        lot_id=lid, lot_number=lot.lot_number, total_products=total_products,
        total_qty=total_qty, parts_total_cost=parts_total, labour_cost=labour,
        total_estimation=total_estimation, notes=notes, grade=grade,
        file_path=stored, file_name=download_name,
        created_by=current_user.username, created_at=stamp,
    )
    db.add(est)
    await db.flush()
    for line in lines:
        db.add(PartEstimateLine(estimate_id=est.id, **line))

    db.add(PartSourcingRequest(
        source="production", part_code=None,
        part_name=f"Request for {lot.lot_number}",
        qty_requested=total_qty, raised_by=current_user.username,
        lot_id=lid, lot_number=lot.lot_number, estimate_id=est.id,
        created_at=stamp,
    ))

    await audit(db, user=current_user, action="PART_ESTIMATE_GENERATED",
                table_name="part_estimates", record_id=str(est.id),
                new_value={"lot": lot.lot_number, "mode": "per-model", "grade": grade,
                           "models": len(priced_blocks), "total_qty": total_qty,
                           "total_estimation": str(total_estimation)},
                request=request)
    await db.commit()

    return JSONResponse({"ok": True, "estimate_id": str(est.id),
                         "download_url": f"/part-estimation/download/{est.id}"})


def _build_workbook(*, lot_number, generated_on, generated_by, lines,
                    total_qty, parts_total, labour_cost, total_estimation, notes) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Part Estimate"

    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money = '#,##0.00'

    ws["A1"] = "Part Estimation"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Lot Number: {lot_number}"
    ws["A2"].font = bold
    ws["A3"] = f"Generated On: {generated_on}"
    ws["A4"] = f"Generated By: {generated_by}"

    header_row = 6
    for col, title in enumerate(["Part Name", "Part Quantity", "Part Unit Cost", "Part Total Cost"], start=1):
        c = ws.cell(row=header_row, column=col, value=title)
        c.fill, c.font, c.border = head_fill, head_font, border
        c.alignment = Alignment(horizontal="center")

    r = header_row
    for line in lines:
        r += 1
        ws.cell(row=r, column=1, value=line["part_name"]).border = border
        ws.cell(row=r, column=2, value=line["qty"]).border = border
        c3 = ws.cell(row=r, column=3, value=float(line["unit_cost"]))
        c3.border, c3.number_format = border, money
        c4 = ws.cell(row=r, column=4, value=float(line["total_cost"]))
        c4.border, c4.number_format = border, money

    r += 1
    ws.cell(row=r, column=1, value="Total").font = bold
    ws.cell(row=r, column=2, value=total_qty).font = bold
    tc = ws.cell(row=r, column=4, value=float(parts_total))
    tc.font, tc.number_format = bold, money
    for col in range(1, 5):
        ws.cell(row=r, column=col).border = border

    r += 2
    ws.cell(row=r, column=1, value="Labour Cost").font = bold
    lc = ws.cell(row=r, column=4, value=float(labour_cost))
    lc.number_format = money
    r += 1
    ws.cell(row=r, column=1, value="Total Estimation").font = bold
    te = ws.cell(row=r, column=4, value=float(total_estimation))
    te.font, te.number_format = bold, money

    if notes:
        r += 2
        ws.cell(row=r, column=1, value="Notes").font = bold
        nc = ws.cell(row=r, column=2, value=notes)
        nc.alignment = Alignment(wrap_text=True, vertical="top")

    for col, width in (("A", 28), ("B", 16), ("C", 18), ("D", 18)):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.post("/{lot_id}/generate")
async def generate_estimate(
    lot_id: str,
    request: Request,
    payload: str = Form(...),          # JSON: {parts:[{part_name,qty,unit_cost}], labour_cost, notes}
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(allowed),
):
    """Download & Attach Estimate — writes the workbook, stores the estimate,
    and raises the production sourcing request in one transaction."""
    import json

    lid = _as_uuid(lot_id)
    lot = (await db.execute(select(Lot).where(Lot.id == lid))).scalar_one_or_none()
    if not lot:
        raise HTTPException(404, "Lot not found")

    try:
        data = json.loads(payload)
    except (ValueError, TypeError):
        return JSONResponse({"ok": False, "error": "Malformed estimate payload"}, status_code=400)

    # Quantities are re-derived from IQC, never taken from the browser — only
    # the unit costs, labour and notes are user input.
    truth = {p["part_name"]: p["qty"] for p in await required_parts_for_lot(db, lid)}
    if not truth:
        return JSONResponse({"ok": False, "error": "No parts are required for this lot"}, status_code=400)

    submitted_costs = {str(p.get("part_name")): p.get("unit_cost") for p in (data.get("parts") or [])}

    lines, total_qty, parts_total = [], 0, Decimal("0")
    for part_name, qty in truth.items():
        unit = _money(submitted_costs.get(part_name, 0))
        total = _money(unit * qty)
        lines.append({"part_name": part_name, "qty": qty, "unit_cost": unit, "total_cost": total})
        total_qty += qty
        parts_total += total

    labour = _money(data.get("labour_cost"))
    total_estimation = _money(parts_total + labour)
    notes = (data.get("notes") or "").strip() or None

    total_products = (await device_counts_by_lot(db)).get(lid, 0)
    stamp = app_now()

    xlsx = _build_workbook(
        lot_number=lot.lot_number, generated_on=stamp.strftime("%d %b %Y %H:%M"),
        generated_by=current_user.username, lines=lines, total_qty=total_qty,
        parts_total=parts_total, labour_cost=labour,
        total_estimation=total_estimation, notes=notes,
    )

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    safe_lot = "".join(ch for ch in (lot.lot_number or "lot") if ch.isalnum() or ch in "-_")
    stored = f"{uuid.uuid4().hex}.xlsx"
    download_name = f"Part_Estimate_{safe_lot}_{stamp.strftime('%Y%m%d-%H%M')}.xlsx"
    with open(os.path.join(UPLOAD_DIR, stored), "wb") as fh:
        fh.write(xlsx)

    est = PartEstimate(
        lot_id=lid, lot_number=lot.lot_number, total_products=total_products,
        total_qty=total_qty, parts_total_cost=parts_total, labour_cost=labour,
        total_estimation=total_estimation, notes=notes,
        file_path=stored, file_name=download_name,
        created_by=current_user.username, created_at=stamp,
    )
    db.add(est)
    await db.flush()

    for line in lines:
        db.add(PartEstimateLine(estimate_id=est.id, **line))

    db.add(PartSourcingRequest(
        source="production", part_code=None,
        part_name=f"Request for {lot.lot_number}",
        qty_requested=total_qty, raised_by=current_user.username,
        lot_id=lid, lot_number=lot.lot_number, estimate_id=est.id,
        created_at=stamp,
    ))

    await audit(db, user=current_user, action="PART_ESTIMATE_GENERATED",
                table_name="part_estimates", record_id=str(est.id),
                new_value={"lot": lot.lot_number, "total_qty": total_qty,
                           "total_estimation": str(total_estimation)},
                request=request)
    await db.commit()

    return JSONResponse({"ok": True, "estimate_id": str(est.id),
                         "download_url": f"/part-estimation/download/{est.id}"})


@router.post("/{estimate_id}/delete-file")
async def delete_estimate_file(
    estimate_id: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(allowed),
):
    """Detach the estimate file and withdraw its Production sourcing request.

    The Download column goes back to empty and the request disappears from
    Part Master -> Sourcing Requests, which is what "Delete File" means to the
    user. Two things are deliberately kept:

    * The workbook itself stays under uploads/part_estimates. Removing the row
      that points at it is reversible; deleting the file is not, and nothing
      else on the page depends on the bytes being gone.
    * The PartEstimate row and its lines stay, so re-running Generate Estimate
      still pre-fills the unit costs and labour that were entered last time.

    Only the attachment fields are cleared, so file_name being empty is what
    hides the button.
    """
    est = (await db.execute(
        select(PartEstimate).where(PartEstimate.id == _as_uuid(estimate_id))
    )).scalar_one_or_none()
    if not est:
        return JSONResponse({"ok": False, "error": "Estimate not found"}, status_code=404)
    if not est.file_name:
        return JSONResponse({"ok": False, "error": "No file is attached to this estimate"},
                            status_code=409)

    srs = (await db.execute(
        select(PartSourcingRequest).where(PartSourcingRequest.estimate_id == est.id)
    )).scalars().all()

    # Snapshot enough to put this back by hand. file_path matters as much as
    # file_name: the stored name is a random hex, so without it the workbook
    # on disk cannot be matched back to the estimate. Same for the sourcing
    # request's own fields — logging only its id makes it unrecoverable.
    undo = {
        "file_name": est.file_name,
        "file_path": est.file_path,
        "sourcing_requests": [{
            "id": str(s.id), "part_name": s.part_name,
            "qty_requested": s.qty_requested, "qty_sourced": s.qty_sourced,
            "raised_by": s.raised_by, "status": s.status,
            "source_deal_id": s.source_deal_id,
            "confirmed": s.confirmed,
            "confirmed_by": s.confirmed_by,
            "confirmed_at": s.confirmed_at.isoformat() if s.confirmed_at else None,
            "closed_by": s.closed_by,
            "closed_at": s.closed_at.isoformat() if s.closed_at else None,
            "created_at": s.created_at.isoformat() if s.created_at else None,
        } for s in srs],
    }
    removed_file = est.file_name

    for sr in srs:
        await db.delete(sr)
    est.file_path = None
    est.file_name = None

    await audit(db, user=current_user, action="PART_ESTIMATE_FILE_REMOVED",
                table_name="part_estimates", record_id=str(est.id),
                old_value=undo,
                new_value={"lot_number": est.lot_number},
                request=request)
    await db.commit()

    msg = f"Estimate file removed for {est.lot_number}"
    if srs:
        msg += f" and {len(srs)} sourcing request(s) withdrawn"
    return JSONResponse({"ok": True, "message": msg})


@router.get("/download/{estimate_id}")
async def download_estimate(
    estimate_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(allowed),
):
    from fastapi.responses import FileResponse

    est = (await db.execute(
        select(PartEstimate).where(PartEstimate.id == _as_uuid(estimate_id))
    )).scalar_one_or_none()
    if not est or not est.file_path:
        raise HTTPException(404, "Estimate file not found")
    path = os.path.join(UPLOAD_DIR, est.file_path)
    if not os.path.exists(path):
        raise HTTPException(404, "Estimate file is no longer on disk")
    return FileResponse(
        path, filename=est.file_name or "part_estimate.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
