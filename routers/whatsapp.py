"""
WhatsApp Router
---------------
Bridges the FastAPI app to the Node.js wa-service (whatsapp-web.js).
wa-service must be running on WA_SERVICE_URL (default http://localhost:3001).

Endpoints:
  GET  /whatsapp              — main page (session, compose, reminders)
  POST /whatsapp/connect      — ask wa-service to start QR flow
  POST /whatsapp/disconnect   — destroy WA session
  GET  /whatsapp/qr-poll      — JSON: { qr_base64 } for live QR display
  GET  /whatsapp/status-poll  — JSON: { status, phone_number, has_qr }
  POST /whatsapp/send         — queue + immediately send a text message
  POST /whatsapp/send-reminder/{order_id}  — payment reminder shortcut
  POST /whatsapp/send-multi-group          — send one message to multiple groups
  POST /whatsapp/groups/{group_id}/tag     — update group tags
  GET  /whatsapp/groups/export-csv         — CSV export of group messages
  GET  /whatsapp/groups/export-xlsx        — Excel export (optional ?ids=id1,id2)
  GET  /whatsapp/groups/messages           — JSON search of group messages
"""
import csv as csv_module
import io as io_module
from templates_config import templates
from datetime import datetime
from utils.timezone import app_now
from fastapi import APIRouter, BackgroundTasks, Depends, Form, Request, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from database import get_db
from models.whatsapp import WhatsAppSession, WhatsAppMessage, WhatsAppGroup, WhatsAppBroadcast
from models.dealers import Dealer, DealerOrder
from models.user import User, UserRole
from auth.dependencies import get_current_user, require_roles, verify_csrf

try:
    import httpx
    _HTTPX_OK = True
except ImportError:
    _HTTPX_OK = False

router = APIRouter(prefix="/whatsapp", tags=["whatsapp"], dependencies=[Depends(verify_csrf)])

WA_SERVICE_URL = "http://localhost:3001"
WA_TIMEOUT     = 8.0   # seconds


# ── Helper: call wa-service ────────────────────────────────────────────────
async def _wa(method: str, path: str, json: dict = None, timeout: float = None, user: str = None):
    """Call the Node wa-service. Returns (status_code, dict) or (0, {}) on error.

    `user` scopes the call to that user's WhatsApp session (per-user multi-session):
    sent as ?user= for GET and injected into the JSON body for POST.
    """
    if not _HTTPX_OK:
        return 0, {"error": "httpx not installed — run: pip install httpx"}
    try:
        t = timeout if timeout is not None else WA_TIMEOUT
        params = {"user": user} if user else None
        async with httpx.AsyncClient(timeout=t) as c:
            if method == "GET":
                r = await c.get(f"{WA_SERVICE_URL}{path}", params=params)
            else:
                body = dict(json or {})
                if user:
                    body.setdefault("user", user)
                r = await c.post(f"{WA_SERVICE_URL}{path}", json=body)
            return r.status_code, r.json()
    except Exception as e:
        return 0, {"error": str(e)}


async def _sync_session(db: AsyncSession, username: str) -> WhatsAppSession:
    """Fetch or create a WhatsAppSession row for this user."""
    result = await db.execute(
        select(WhatsAppSession).where(WhatsAppSession.username == username)
    )
    session = result.scalar_one_or_none()
    if not session:
        session = WhatsAppSession(username=username, status="disconnected")
        db.add(session)
        await db.flush()
    return session


# ── Compose / pre-fill page ────────────────────────────────────────────────
@router.get("/compose", response_class=HTMLResponse)
async def compose(
    request: Request,
    order_id: str = Query(default=""),
    type: str = Query(default=""),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Pre-fill the WhatsApp send form from a dealer order."""
    if not order_id:
        return RedirectResponse(url="/whatsapp?error=No+order+specified", status_code=302)

    order_result = await db.execute(
        select(DealerOrder).where(DealerOrder.id == order_id)
    )
    order = order_result.scalar_one_or_none()
    if order is None:
        return RedirectResponse(url="/whatsapp?error=Order+not+found", status_code=302)

    dealer_result = await db.execute(
        select(Dealer).where(Dealer.id == order.dealer_id)
    )
    dealer = dealer_result.scalar_one_or_none()
    if dealer is None:
        return RedirectResponse(url="/whatsapp?error=Dealer+not+found", status_code=302)

    prefill_phone = dealer.whatsapp_number or dealer.phone or ""
    greeting_name = dealer.first_name or dealer.business_name
    prefill_message = (
        f"Dear {greeting_name},\n\n"
        f"This is a payment reminder for invoice "
        f"{order.invoice_number or order.order_number}.\n"
        f"Outstanding amount: ₹{int(order.due_amount):,}\n\n"
        f"Please arrange payment at your earliest convenience."
        f"\n\nRegards,\nOxyPC Team"
    )

    return templates.TemplateResponse("whatsapp/compose.html", {
        "request": request,
        "current_user": current_user,
        "dealer": dealer,
        "order": order,
        "prefill_phone": prefill_phone,
        "prefill_message": prefill_message,
    })


# ── Main page ──────────────────────────────────────────────────────────────
@router.get("", response_class=HTMLResponse)
async def index(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    session = await _sync_session(db, current_user.username)

    # Sync status from wa-service (non-blocking: if service offline, keep DB value)
    _, wa_data = await _wa("GET", "/status", user=current_user.username)
    if wa_data.get("status") in ("connected", "scanning", "disconnected"):
        new_status = wa_data["status"]
        if session.status != new_status:
            session.status = new_status
        if wa_data.get("phone_number"):
            session.phone_number = wa_data["phone_number"]
        if new_status == "connected" and not session.connected_at:
            session.connected_at = app_now()
        elif new_status == "disconnected":
            session.connected_at = None
    await db.commit()

    # Recent messages
    msgs_result = await db.execute(
        select(WhatsAppMessage)
        .where(WhatsAppMessage.sent_by == current_user.username)
        .order_by(WhatsAppMessage.created_at.desc())
        .limit(15)
    )
    recent_messages = msgs_result.scalars().all()

    # Active dealers for compose
    dealers_result = await db.execute(
        select(Dealer).where(Dealer.status == "active").order_by(Dealer.business_name)
    )
    dealers = dealers_result.scalars().all()

    # Groups
    groups_result = await db.execute(
        select(WhatsAppGroup).order_by(WhatsAppGroup.group_name)
    )
    wa_groups = groups_result.scalars().all()

    # Recent broadcasts
    bc_result = await db.execute(
        select(WhatsAppBroadcast)
        .where(WhatsAppBroadcast.sent_by == current_user.username)
        .order_by(WhatsAppBroadcast.created_at.desc())
        .limit(10)
    )
    broadcasts = bc_result.scalars().all()

    # Payment reminders due
    reminders_result = await db.execute(
        select(DealerOrder, Dealer.business_name)
        .join(Dealer, DealerOrder.dealer_id == Dealer.id)
        .where(DealerOrder.due_amount > 0, DealerOrder.status != "cancelled")
        .order_by(DealerOrder.payment_due_date.asc().nullslast())
        .limit(20)
    )
    payment_reminders = []
    for order, dealer_name in reminders_result.all():
        payment_reminders.append({
            "id":               order.id,
            "order_number":     order.order_number,
            "invoice_number":   order.invoice_number,
            "dealer_name":      dealer_name,
            "due_amount":       float(order.due_amount or 0),
            "payment_due_date": order.payment_due_date,
        })

    service_online = wa_data.get("error") is None

    return templates.TemplateResponse("whatsapp/index.html", {
        "request":           request,
        "current_user":      current_user,
        "session":           session,
        "recent_messages":   recent_messages,
        "dealers":           dealers,
        "wa_groups":         wa_groups,
        "broadcasts":        broadcasts,
        "payment_reminders": payment_reminders,
        "today":             app_now().date(),
        "service_online":    service_online,
    })


# ── Connect ────────────────────────────────────────────────────────────────
@router.post("/connect")
async def connect(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    session = await _sync_session(db, current_user.username)
    _, data  = await _wa("POST", "/connect", user=current_user.username)

    session.status = data.get("status", "scanning")
    await db.commit()

    if data.get("error"):
        return RedirectResponse(
            url=f"/whatsapp?error=WA+service+offline%3A+{data['error'][:60]}",
            status_code=302
        )
    return RedirectResponse(url="/whatsapp?success=Scan+QR+code+to+connect", status_code=302)


# ── Disconnect ─────────────────────────────────────────────────────────────
@router.post("/disconnect")
async def disconnect(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    session = await _sync_session(db, current_user.username)
    await _wa("POST", "/disconnect", user=current_user.username)

    session.status       = "disconnected"
    session.session_data = None
    session.connected_at = None
    await db.commit()
    return RedirectResponse(url="/whatsapp?success=Disconnected", status_code=302)


# ── Live QR poll (called by JS every 3 s while scanning) ──────────────────
@router.get("/qr-poll")
async def qr_poll(
    request: Request,
    current_user: User = Depends(get_current_user),
):
    code, data = await _wa("GET", "/qr", user=current_user.username)
    if code == 200 and data.get("qr_base64"):
        return JSONResponse({"qr_base64": data["qr_base64"]})
    return JSONResponse({"error": "no_qr"}, status_code=404)


# ── Status poll (called by JS to detect scan completion) ──────────────────
@router.get("/status-poll")
async def status_poll(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _, data = await _wa("GET", "/status", user=current_user.username)
    if data.get("status") in ("connected", "scanning", "disconnected"):
        session = await _sync_session(db, current_user.username)
        session.status = data["status"]
        if data.get("phone_number"):
            session.phone_number = data["phone_number"]
        if data["status"] == "connected" and not session.connected_at:
            session.connected_at = app_now()
        await db.commit()
    return JSONResponse({
        "status":       data.get("status", "unknown"),
        "phone_number": data.get("phone_number"),
        "has_qr":       data.get("has_qr", False),
        "service_online": data.get("error") is None,
    })


# ── Send message ───────────────────────────────────────────────────────────
@router.post("/send")
async def send_message(
    request: Request,
    dealer_id:       str = Form(default=None),
    recipient_phone: str = Form(...),
    message_type:    str = Form(default="text"),
    message_text:    str = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    dealer_name = None
    if dealer_id:
        d = (await db.execute(select(Dealer).where(Dealer.id == dealer_id))).scalar_one_or_none()
        if d:
            dealer_name = d.business_name

    msg = WhatsAppMessage(
        sent_by         = current_user.username,
        recipient_phone = recipient_phone,
        recipient_name  = dealer_name,
        message_type    = message_type,
        message_text    = message_text,
        dealer_id       = dealer_id or None,
        status          = "pending",
    )
    db.add(msg)

    # Try immediate send via wa-service
    session = await _sync_session(db, current_user.username)
    if session.status == "connected":
        code, resp = await _wa("POST", "/send", json={
            "phone":   recipient_phone,
            "message": message_text,
        }, user=current_user.username)
        if code == 200 and resp.get("success"):
            msg.status  = "sent"
            msg.sent_at = app_now()
        else:
            msg.status    = "failed"
            msg.error_msg = resp.get("error", "Send error")

    await db.commit()

    if msg.status == "sent":
        return RedirectResponse(url="/whatsapp?success=Message+sent!", status_code=302)
    elif msg.status == "failed":
        return RedirectResponse(url=f"/whatsapp?error=Send+failed:+{msg.error_msg[:60]}", status_code=302)
    return RedirectResponse(url="/whatsapp?success=Message+queued", status_code=302)


# ── Payment reminder shortcut ──────────────────────────────────────────────
@router.post("/send-reminder/{order_id}")
async def send_payment_reminder(
    request:  Request,
    order_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    order = (await db.execute(select(DealerOrder).where(DealerOrder.id == order_id))).scalar_one_or_none()
    if not order:
        return RedirectResponse(url="/whatsapp?error=Order+not+found", status_code=302)

    dealer = (await db.execute(select(Dealer).where(Dealer.id == order.dealer_id))).scalar_one_or_none()
    phone  = (dealer.whatsapp_number or dealer.phone) if dealer else None
    if not phone:
        return RedirectResponse(url="/whatsapp?error=No+WhatsApp+number+for+dealer", status_code=302)

    greeting_name = (dealer.first_name or dealer.business_name) if dealer else "Customer"
    text = (
        f"Dear {greeting_name},\n\n"
        f"This is a payment reminder for invoice {order.invoice_number or order.order_number}.\n"
        f"Outstanding amount: \u20b9{int(order.due_amount):,}\n\n"
        f"Please arrange payment at your earliest convenience.\n\nRegards,\nOxyPC Team"
    )

    msg = WhatsAppMessage(
        sent_by         = current_user.username,
        recipient_phone = phone,
        recipient_name  = dealer.business_name if dealer else None,
        message_type    = "reminder",
        message_text    = text,
        dealer_id       = str(order.dealer_id) if order.dealer_id else None,
        reference_type  = "payment_reminder",
        reference_id    = str(order.id),
        status          = "pending",
    )
    db.add(msg)
    order.payment_reminder_sent = True

    # Try immediate send
    session = await _sync_session(db, current_user.username)
    if session.status == "connected":
        code, resp = await _wa("POST", "/send", json={"phone": phone, "message": text}, user=current_user.username)
        if code == 200 and resp.get("success"):
            msg.status  = "sent"
            msg.sent_at = app_now()
        else:
            msg.status    = "failed"
            msg.error_msg = resp.get("error", "")

    await db.commit()
    return RedirectResponse(url="/whatsapp?success=Payment+reminder+sent", status_code=302)


# ── Incoming group message webhook (called by wa-service) ─────────────────
@router.post("/incoming-group-msg")
async def incoming_group_message(
    request: Request,
    db: AsyncSession = Depends(get_db),
):
    # Simple shared-secret guard — wa-service sends this header
    if request.headers.get("X-WA-Secret") != "oxypc-wa-internal":
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Forbidden")

    data = await request.json()
    group_id   = data.get("group_id", "")
    group_name = data.get("group_name", group_id)
    msg_text   = data.get("message_text", "")
    if not group_id or not msg_text:
        return JSONResponse({"ok": False})

    # Check for duplicate (same group + same timestamp)
    ts = data.get("timestamp")
    if ts:
        from sqlalchemy import and_
        dup = (await db.execute(
            select(WhatsAppMessage).where(
                and_(
                    WhatsAppMessage.recipient_phone == group_id,
                    WhatsAppMessage.sender_phone == data.get("sender_phone", ""),
                    WhatsAppMessage.message_text == msg_text,
                    WhatsAppMessage.direction == "incoming",
                )
            ).limit(1)
        )).scalar_one_or_none()
        if dup:
            return JSONResponse({"ok": True, "dup": True})

    msg = WhatsAppMessage(
        sent_by        = "incoming",
        recipient_phone= group_id,
        recipient_name = group_name,
        message_type   = data.get("message_type", "text"),
        message_text   = msg_text,
        reference_type = "group",
        direction      = "incoming",
        sender_name    = data.get("sender_name", ""),
        sender_phone   = data.get("sender_phone", ""),
        status         = "received",
        sent_at        = datetime.utcfromtimestamp(ts) if ts else app_now(),
    )
    db.add(msg)
    await db.commit()
    return JSONResponse({"ok": True})


# ── In-memory sync status ─────────────────────────────────────────────────
_sync_status: dict = {"running": False, "saved": 0, "skipped": 0, "errors": 0, "done": False, "message": ""}


async def _do_sync_messages(group_ids: list, limit: int, username: str):
    """Background task: fetch messages from all groups and store in DB."""
    from sqlalchemy import and_
    from database import AsyncSessionLocal  # direct session, not request-scoped
    _sync_status.update({"running": True, "saved": 0, "skipped": 0, "errors": 0, "done": False, "message": "Syncing…"})
    total_saved = total_skipped = 0
    BATCH = 10   # smaller batches so each wa-service call is faster
    async with AsyncSessionLocal() as db:
        for i in range(0, len(group_ids), BATCH):
            batch = group_ids[i:i+BATCH]
            code, data = await _wa("POST", "/sync-group-messages", json={"group_ids": batch, "limit": limit}, timeout=120, user=username)
            if code != 200:
                _sync_status["errors"] += len(batch)
                err_msg = data.get("error", "")
                if "not connected" in err_msg.lower() or "reconnect" in err_msg.lower():
                    _sync_status.update({"message": "WA session lost — wa-service is reconnecting. Sync paused.", "running": False, "done": True})
                    break
                continue
            # If wa-service detected a broken session mid-batch, stop here
            if data.get("session_broken"):
                _sync_status["errors"] += 1
                _sync_status.update({"message": f"WA session reset mid-sync after {min(i+BATCH, len(group_ids))}/{len(group_ids)} groups. Saved {total_saved} messages. Restart wa-service and sync again to continue."})
                # Still process whatever messages came back before the break
            for m in data.get("messages", []):
                msg_text = m.get("message_text", "")
                if not msg_text:
                    continue
                gid = m.get("group_id", "")
                ts  = m.get("timestamp")
                dup = (await db.execute(
                    select(WhatsAppMessage.id).where(
                        and_(
                            WhatsAppMessage.recipient_phone == gid,
                            WhatsAppMessage.message_text   == msg_text,
                            WhatsAppMessage.sender_phone   == (m.get("sender_phone") or ""),
                            WhatsAppMessage.direction      == ("outgoing" if m.get("from_me") else "incoming"),
                        )
                    ).limit(1)
                )).scalar_one_or_none()
                if dup:
                    total_skipped += 1
                    continue
                msg = WhatsAppMessage(
                    sent_by        = username if m.get("from_me") else "incoming",
                    recipient_phone= gid,
                    recipient_name = m.get("group_name", gid),
                    message_type   = m.get("message_type", "text"),
                    message_text   = msg_text[:4000],
                    reference_type = "group",
                    direction      = "outgoing" if m.get("from_me") else "incoming",
                    sender_name    = (m.get("sender_name", "") or "")[:200],
                    sender_phone   = (m.get("sender_phone", "") or "")[:30],
                    status         = "sent" if m.get("from_me") else "received",
                    sent_at        = datetime.utcfromtimestamp(ts) if ts else app_now(),
                )
                db.add(msg)
                total_saved += 1
            await db.commit()
            _sync_status.update({"saved": total_saved, "skipped": total_skipped,
                                  "message": f"Processed {min(i+BATCH, len(group_ids))}/{len(group_ids)} groups…"})

    # Only overwrite message if it wasn't already set by a session-broken early-exit
    if _sync_status.get("running"):
        _sync_status.update({
            "running": False, "done": True,
            "message": f"Done — {total_saved} new messages saved, {total_skipped} already existed."
        })
    else:
        # session_broken or other early exit already set the message
        _sync_status.update({"running": False, "done": True, "saved": total_saved, "skipped": total_skipped})


# ── Bulk sync messages from WA groups ──────────────────────────────────────
@router.post("/groups/sync-messages")
async def sync_group_messages(
    request: Request,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if _sync_status["running"]:
        return RedirectResponse(url="/whatsapp?tab=groups&info=Sync+already+running+in+background", status_code=302)

    form      = await request.form()
    group_ids = form.getlist("group_ids")   # empty = sync all groups
    limit     = int(form.get("limit", "50"))

    if not group_ids:
        result = await db.execute(select(WhatsAppGroup).order_by(WhatsAppGroup.group_name))
        group_ids = [g.group_wa_id for g in result.scalars().all()]

    if not group_ids:
        return RedirectResponse(url="/whatsapp?tab=groups&error=No+groups+to+sync", status_code=302)

    background_tasks.add_task(_do_sync_messages, group_ids, limit, current_user.username)
    n = len(group_ids)
    return RedirectResponse(
        url=f"/whatsapp?tab=groups&success=Sync+started+for+{n}+groups+in+background.+Come+back+in+a+minute+and+refresh.",
        status_code=302,
    )


# ── Sync status poll ────────────────────────────────────────────────────────
@router.get("/groups/sync-status")
async def sync_status_poll(current_user: User = Depends(get_current_user)):
    return JSONResponse(_sync_status)


# ── Sync groups from wa-service ────────────────────────────────────────────
@router.post("/groups/sync")
async def sync_groups(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    code, data = await _wa("GET", "/groups", timeout=30, user=current_user.username)
    if code != 200 or "error" in data:
        err_raw = data.get("error", "WA offline")
        # Detect Puppeteer session errors — show friendly reconnect message
        _broken_keywords = ("detached", "execution context", "session closed",
                            "target closed", "protocol error", "attempted to use")
        if any(k in err_raw.lower() for k in _broken_keywords) or code == 503:
            return RedirectResponse(
                url="/whatsapp?tab=groups&error=WA+session+reset+—+wait+15+seconds+then+click+Sync+Groups+again",
                status_code=302,
            )
        return RedirectResponse(
            url=f"/whatsapp?tab=groups&error=Groups+sync+failed:+{err_raw[:60]}",
            status_code=302,
        )
    groups = data.get("groups", [])
    synced = 0
    for g in groups:
        existing = (await db.execute(
            select(WhatsAppGroup).where(WhatsAppGroup.group_wa_id == g["id"])
        )).scalar_one_or_none()
        if existing:
            existing.group_name        = g["name"]
            existing.participant_count = g.get("participant_count", 0)
            existing.last_synced       = app_now()
            existing.synced_by         = current_user.username
        else:
            db.add(WhatsAppGroup(
                group_wa_id       = g["id"],
                group_name        = g["name"],
                participant_count = g.get("participant_count", 0),
                synced_by         = current_user.username,
            ))
        synced += 1
    await db.commit()
    return RedirectResponse(
        url=f"/whatsapp?tab=groups&success=Synced+{synced}+groups", status_code=302
    )


# ── Send message to a group ────────────────────────────────────────────────
@router.post("/send-group")
async def send_group_message(
    request: Request,
    group_wa_id:  str = Form(...),
    group_name:   str = Form(default=""),
    message_type: str = Form(default="text"),
    message_text: str = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Log as a WhatsAppMessage (recipient_phone = group WA id)
    msg = WhatsAppMessage(
        sent_by         = current_user.username,
        recipient_phone = group_wa_id,
        recipient_name  = group_name or group_wa_id,
        message_type    = message_type,
        message_text    = message_text,
        reference_type  = "group",
        status          = "pending",
    )
    db.add(msg)

    session = await _sync_session(db, current_user.username)
    if session.status == "connected":
        code, resp = await _wa("POST", "/send-group", json={
            "group_id": group_wa_id,
            "message":  message_text,
        }, user=current_user.username)
        if code == 200 and resp.get("success"):
            msg.status  = "sent"
            msg.sent_at = app_now()
        elif code == 503:
            # WA session lost — auto-reconnect triggered on wa-service side
            msg.status    = "failed"
            msg.error_msg = "WA session lost (auto-reconnecting)"
            await db.commit()
            return RedirectResponse(
                url="/whatsapp?tab=groups&error=WA+session+lost+—+reconnecting+automatically.+Wait+15+seconds+then+try+again.",
                status_code=302)
        else:
            msg.status    = "failed"
            msg.error_msg = resp.get("error", "Group send error")

    await db.commit()
    if msg.status == "sent":
        return RedirectResponse(url="/whatsapp?tab=groups&success=Group+message+sent!", status_code=302)
    return RedirectResponse(url=f"/whatsapp?tab=groups&error=Send+failed:+{msg.error_msg[:50]}", status_code=302)


# ── Broadcast to multiple dealers ──────────────────────────────────────────
@router.post("/broadcast")
async def broadcast_message(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    form       = await request.form()
    dealer_ids = form.getlist("dealer_ids")
    msg_type   = form.get("message_type", "text")
    msg_text   = form.get("message_text", "")
    bc_name    = form.get("broadcast_name", "")

    if not dealer_ids or not msg_text:
        return RedirectResponse(url="/whatsapp?tab=broadcast&error=Select+dealers+and+enter+message", status_code=302)

    # Load dealers
    dealers_result = await db.execute(
        select(Dealer).where(Dealer.id.in_(dealer_ids))
    )
    dealers_list = dealers_result.scalars().all()

    session  = await _sync_session(db, current_user.username)
    sent     = 0
    failed   = 0

    bc = WhatsAppBroadcast(
        broadcast_name   = bc_name or f"Broadcast {app_now().strftime('%d-%m %H:%M')}",
        message_type     = msg_type,
        message_text     = msg_text,
        sent_by          = current_user.username,
        total_recipients = len(dealers_list),
    )
    db.add(bc)
    await db.flush()

    for dealer in dealers_list:
        phone = dealer.whatsapp_number or dealer.phone
        if not phone:
            failed += 1
            continue
        greeting_name = dealer.first_name or dealer.business_name
        personalized = msg_text.replace("{name}", greeting_name)
        msg = WhatsAppMessage(
            sent_by         = current_user.username,
            recipient_phone = phone,
            recipient_name  = dealer.business_name,
            message_type    = msg_type,
            message_text    = personalized,
            dealer_id       = dealer.id,
            reference_type  = "broadcast",
            reference_id    = str(bc.id),
            status          = "pending",
        )
        db.add(msg)
        if session.status == "connected":
            code, resp = await _wa("POST", "/send", json={"phone": phone, "message": personalized}, user=current_user.username)
            if code == 200 and resp.get("success"):
                msg.status  = "sent"
                msg.sent_at = app_now()
                sent += 1
            else:
                msg.status    = "failed"
                msg.error_msg = resp.get("error", "")
                failed += 1
        else:
            failed += 1

    bc.sent_count   = sent
    bc.failed_count = failed
    bc.status       = "done" if failed == 0 else ("partial" if sent > 0 else "failed")
    await db.commit()

    return RedirectResponse(
        url=f"/whatsapp?tab=broadcast&success=Broadcast+sent:+{sent}+ok,+{failed}+failed",
        status_code=302,
    )


# ── Send one message to multiple groups ───────────────────────────────────
@router.post("/send-multi-group")
async def send_multi_group(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    form         = await request.form()
    group_ids    = form.getlist("group_ids")
    msg_type     = form.get("message_type", "text")
    msg_text     = form.get("message_text", "")

    if not group_ids or not msg_text:
        return RedirectResponse(url="/whatsapp?tab=groups&error=Select+groups+and+enter+message", status_code=302)

    session = await _sync_session(db, current_user.username)
    sent, failed = 0, 0

    for gid in group_ids:
        group = (await db.execute(
            select(WhatsAppGroup).where(WhatsAppGroup.group_wa_id == gid)
        )).scalar_one_or_none()
        group_name = group.group_name if group else gid

        msg = WhatsAppMessage(
            sent_by         = current_user.username,
            recipient_phone = gid,
            recipient_name  = group_name,
            message_type    = msg_type,
            message_text    = msg_text,
            reference_type  = "group",
            status          = "pending",
        )
        db.add(msg)

        if session.status == "connected":
            code, resp = await _wa("POST", "/send-group", json={"group_id": gid, "message": msg_text}, user=current_user.username)
            if code == 200 and resp.get("success"):
                msg.status  = "sent"
                msg.sent_at = app_now()
                sent += 1
            elif code == 503:
                # WA session broken — stop sending remaining groups
                msg.status    = "failed"
                msg.error_msg = "WA session lost (auto-reconnecting)"
                failed += 1
                await db.commit()
                return RedirectResponse(
                    url=f"/whatsapp?tab=groups&error=WA+session+lost+after+{sent}+sends+—+reconnecting.+Wait+15s+then+retry.",
                    status_code=302)
            else:
                msg.status    = "failed"
                msg.error_msg = resp.get("error", "Group send error")
                failed += 1
        else:
            failed += 1

    await db.commit()
    return RedirectResponse(
        url=f"/whatsapp?tab=groups&success=Sent+to+{sent}+groups,+{failed}+failed",
        status_code=302,
    )


# ── Update group tags ──────────────────────────────────────────────────────
@router.post("/groups/{group_wa_id_b64}/tag")
async def update_group_tags(
    request: Request,
    group_wa_id_b64: str,
    tags: str = Form(default=""),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # group_wa_id may contain special chars — we pass id via form instead
    group_id_raw = (await request.form()).get("group_wa_id", group_wa_id_b64)
    group = (await db.execute(
        select(WhatsAppGroup).where(WhatsAppGroup.group_wa_id == group_id_raw)
    )).scalar_one_or_none()
    if group:
        group.tags = tags.strip() or None
        await db.commit()
    return RedirectResponse(url="/whatsapp?tab=groups&success=Tags+updated", status_code=302)


# ── Export group messages as CSV ──────────────────────────────────────────
@router.get("/groups/export-csv")
async def export_group_messages_csv(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(WhatsAppMessage)
        .where(WhatsAppMessage.reference_type == "group")
        .order_by(WhatsAppMessage.created_at.desc())
        .limit(2000)
    )
    msgs = result.scalars().all()

    output = io_module.StringIO()
    writer = csv_module.writer(output)
    writer.writerow(["Group Name", "Message Type", "Message Text", "Status", "Sent At", "Sent By"])
    for m in msgs:
        writer.writerow([
            m.recipient_name or m.recipient_phone,
            m.message_type,
            m.message_text or "",
            m.status,
            m.sent_at.strftime("%Y-%m-%d %H:%M") if m.sent_at else "",
            m.sent_by,
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=group_messages.csv"},
    )


# ── Export group messages as Excel (.xlsx) ───────────────────────────────
@router.get("/groups/export-xlsx")
async def export_group_messages_xlsx(
    request: Request,
    ids: str = Query(default=""),          # optional comma-separated message IDs
    q:   str = Query(default=""),          # keyword search filter
    group_id: str = Query(default=""),     # filter by group_wa_id
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from openpyxl import Workbook
    from sqlalchemy import or_

    query = select(WhatsAppMessage).where(WhatsAppMessage.reference_type == "group")

    # Filter by selected IDs if provided
    if ids.strip():
        id_list = [i.strip() for i in ids.split(",") if i.strip()]
        if id_list:
            query = query.where(WhatsAppMessage.id.in_(id_list))
    else:
        # Apply search / group filter
        if q.strip():
            query = query.where(WhatsAppMessage.message_text.ilike(f"%{q.strip()}%"))
        if group_id.strip():
            query = query.where(WhatsAppMessage.recipient_phone == group_id.strip())

    result = await db.execute(query.order_by(WhatsAppMessage.created_at.desc()).limit(2000))
    msgs = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Group Messages"

    # Header row
    headers = ["#", "Group Name", "Message Type", "Message Text", "Status", "Sent At", "Sent By"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = __import__("openpyxl").styles.Font(bold=True)
        cell.fill = __import__("openpyxl").styles.PatternFill("solid", fgColor="075E54")
        cell.font = __import__("openpyxl").styles.Font(bold=True, color="FFFFFF")

    for i, m in enumerate(msgs, 1):
        ws.append([
            i,
            m.recipient_name or m.recipient_phone,
            m.message_type or "text",
            m.message_text or "",
            m.status,
            m.sent_at.strftime("%Y-%m-%d %H:%M") if m.sent_at else "",
            m.sent_by,
        ])

    # Column widths
    for col, width in zip(["A","B","C","D","E","F","G"], [5, 28, 14, 60, 10, 18, 14]):
        ws.column_dimensions[col].width = width

    output = io_module.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=group_messages.xlsx"},
    )


# ── Group messages JSON search (AJAX) ────────────────────────────────────
@router.get("/groups/messages")
async def group_messages_search(
    request: Request,
    q:        str = Query(default=""),
    group_id: str = Query(default=""),
    page:     int = Query(default=1),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from sqlalchemy import func as sa_func
    PAGE_SIZE = 50
    query = select(WhatsAppMessage).where(WhatsAppMessage.reference_type == "group")

    if q.strip():
        query = query.where(WhatsAppMessage.message_text.ilike(f"%{q.strip()}%"))
    if group_id.strip():
        query = query.where(WhatsAppMessage.recipient_phone == group_id.strip())

    count_result = await db.execute(select(sa_func.count()).select_from(query.subquery()))
    total = count_result.scalar() or 0

    offset = (page - 1) * PAGE_SIZE
    result = await db.execute(
        query.order_by(WhatsAppMessage.created_at.desc()).offset(offset).limit(PAGE_SIZE)
    )
    msgs = result.scalars().all()

    return JSONResponse({
        "total": total,
        "page": page,
        "pages": max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE),
        "messages": [{
            "id":           str(m.id),
            "group_name":   m.recipient_name or m.recipient_phone,
            "group_id":     m.recipient_phone,
            "message_type": m.message_type or "text",
            "message_text": m.message_text or "",
            "status":       m.status,
            "direction":    getattr(m, "direction", "outgoing") or "outgoing",
            "sender_name":  getattr(m, "sender_name", "") or "",
            "sender_phone": getattr(m, "sender_phone", "") or "",
            "sent_at":      m.sent_at.strftime("%d-%m-%Y %H:%M") if m.sent_at else None,
            "created_at":   m.created_at.strftime("%d-%m-%Y %H:%M") if m.created_at else None,
            "sent_by":      m.sent_by,
        } for m in msgs],
    })


# ── Group message analytics (sender frequency) ────────────────────────────
@router.get("/groups/analytics")
async def group_analytics(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from sqlalchemy import func as sa_func
    # Top senders by message count
    rows = (await db.execute(
        select(
            WhatsAppMessage.sender_phone,
            WhatsAppMessage.sender_name,
            sa_func.count().label("msg_count"),
            sa_func.count(WhatsAppMessage.recipient_phone.distinct()).label("group_count"),
        )
        .where(WhatsAppMessage.reference_type == "group")
        .where(WhatsAppMessage.direction      == "incoming")
        .where(WhatsAppMessage.sender_phone   != "")
        .where(WhatsAppMessage.sender_phone   != None)
        .group_by(WhatsAppMessage.sender_phone, WhatsAppMessage.sender_name)
        .order_by(sa_func.count().desc())
        .limit(200)
    )).all()

    return JSONResponse([{
        "phone":       r.sender_phone,
        "name":        r.sender_name or "",
        "msg_count":   r.msg_count,
        "group_count": r.group_count,
    } for r in rows])


# ── Add WA sender as dealer ────────────────────────────────────────────────
@router.post("/groups/add-sender-dealer")
async def add_sender_as_dealer(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    form         = await request.form()
    phone        = (form.get("phone") or "").strip()
    name         = (form.get("name") or "").strip()
    business     = (form.get("business_name") or name or phone).strip()
    source_group = (form.get("source_group") or "").strip()

    if not phone:
        return JSONResponse({"ok": False, "error": "phone required"}, status_code=400)

    # Check if dealer with this phone already exists
    existing = (await db.execute(
        select(Dealer).where(Dealer.phone == phone)
    )).scalar_one_or_none()

    if existing:
        return JSONResponse({"ok": False, "error": f"Dealer already exists: {existing.business_name}", "dealer_id": str(existing.id)})

    # Split name into first/last
    parts = name.split(" ", 1)
    new_dealer = Dealer(
        business_name = business,
        first_name    = parts[0] if parts else "",
        last_name     = parts[1] if len(parts) > 1 else "",
        phone         = phone,
        source        = f"WhatsApp Group{': ' + source_group if source_group else ''}",
        added_by      = current_user.username,
    )
    db.add(new_dealer)
    await db.commit()
    await db.refresh(new_dealer)
    return JSONResponse({"ok": True, "dealer_id": str(new_dealer.id), "business_name": new_dealer.business_name})


# ── Admin: All WA messages audit log ─────────────────────────────────────
_admin_only = require_roles(UserRole.admin)

@router.get("/audit", response_class=HTMLResponse)
async def wa_audit_log(
    request:      Request,
    q:            str = Query(default=""),
    user:         str = Query(default=""),
    msg_type:     str = Query(default=""),
    status:       str = Query(default=""),
    page:         int = Query(default=1),
    db:           AsyncSession = Depends(get_db),
    current_user: User = Depends(_admin_only),
):
    from sqlalchemy import or_
    PAGE_SIZE = 50
    query = select(WhatsAppMessage)
    if q:
        like = f"%{q}%"
        query = query.where(or_(
            WhatsAppMessage.recipient_name.ilike(like),
            WhatsAppMessage.recipient_phone.ilike(like),
            WhatsAppMessage.message_text.ilike(like),
        ))
    if user:
        query = query.where(WhatsAppMessage.sent_by == user)
    if msg_type:
        query = query.where(WhatsAppMessage.message_type == msg_type)
    if status:
        query = query.where(WhatsAppMessage.status == status)

    # Count
    from sqlalchemy import func as sa_func
    count_result = await db.execute(
        select(sa_func.count()).select_from(query.subquery())
    )
    total = count_result.scalar() or 0

    offset = (page - 1) * PAGE_SIZE
    msgs_result = await db.execute(
        query.order_by(WhatsAppMessage.created_at.desc())
        .offset(offset).limit(PAGE_SIZE)
    )
    msgs = msgs_result.scalars().all()

    # Distinct senders for filter
    users_result = await db.execute(
        select(WhatsAppMessage.sent_by).distinct()
    )
    all_users = [r[0] for r in users_result.fetchall()]

    return templates.TemplateResponse("whatsapp/audit.html", {
        "request":      request,
        "current_user": current_user,
        "msgs":         msgs,
        "q":            q,
        "user":         user,
        "msg_type":     msg_type,
        "status":       status,
        "page":         page,
        "total":        total,
        "pages":        (total + PAGE_SIZE - 1) // PAGE_SIZE,
        "all_users":    all_users,
    })


# ── Dealer WA message history (JSON for dealer profile AJAX) ─────────────
@router.get("/dealer/{dealer_id}/history")
async def dealer_wa_history(
    request:   Request,
    dealer_id: str,
    db:        AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(WhatsAppMessage)
        .where(WhatsAppMessage.dealer_id == dealer_id)
        .order_by(WhatsAppMessage.created_at.desc())
        .limit(50)
    )
    msgs = result.scalars().all()
    return JSONResponse([{
        "id":               str(m.id),
        "message_type":     m.message_type,
        "message_text":     m.message_text,
        "status":           m.status,
        "sent_at":          m.sent_at.strftime("%d-%m-%Y %H:%M") if m.sent_at else None,
        "created_at":       m.created_at.strftime("%d-%m-%Y %H:%M"),
        "reference_type":   m.reference_type,
        "sent_by":          m.sent_by,
    } for m in msgs])
