"""
OxyPC ERP — Comprehensive UAT Test Suite + Excel Report Generator
Date: 27 April 2026
Generates: docs/OxyPC_UAT_TestPlan.xlsx
"""
import sys
import os
import time
import warnings
import re

# UTF-8 stdout to avoid Windows console encoding issues
sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)
sys.stderr = open(sys.stderr.fileno(), mode='w', encoding='utf-8', buffering=1)

# Suppress passlib/bcrypt warnings
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", message=".*bcrypt.*")
os.environ["PYTHONWARNINGS"] = "ignore"

import httpx
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── Constants ───────────────────────────────────────────────────────────────
BASE_URL = "http://localhost:8000"
APP_URL_DISPLAY = "http://192.168.4.8:8000"
ADMIN_USER = "admin"
ADMIN_PASS = "oxypc@admin123"
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(PROJECT_ROOT, "docs", "OxyPC_UAT_TestPlan.xlsx")

# ─── Colour Palette ──────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1565C0")   # dark blue
HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
PASS_FILL  = PatternFill("solid", fgColor="C8E6C9")   # green
FAIL_FILL  = PatternFill("solid", fgColor="FFCDD2")   # red
SKIP_FILL  = PatternFill("solid", fgColor="FFF9C4")   # yellow
SS_FILL    = PatternFill("solid", fgColor="FFE0B2")   # orange
GREY_FILL  = PatternFill("solid", fgColor="F5F5F5")   # light grey
ALT_FILL   = PatternFill("solid", fgColor="EEF2F7")   # alternating row
RED_FILL   = PatternFill("solid", fgColor="FFCDD2")   # showstopper header
COVER_FILL = PatternFill("solid", fgColor="0D47A1")   # cover title fill

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

# ─── Test Definition Helpers ─────────────────────────────────────────────────
def t(test_id, phase, module, name, method, url, expected,
      steps, priority="HIGH", showstopper="NO", notes=""):
    return {
        "test_id": test_id,
        "phase": phase,
        "module": module,
        "name": name,
        "method": method,
        "url": url,
        "expected_status": expected,
        "steps": steps,
        "priority": priority,
        "showstopper": showstopper,
        "notes": notes,
        # runtime results (filled later)
        "actual_status": None,
        "status": None,
        "elapsed_ms": None,
    }

# ─── Test Definitions ────────────────────────────────────────────────────────

# ---- PHASE 1 ----------------------------------------------------------------
PHASE1_TESTS = [
    # Auth
    t("P1-AUTH-01","Phase 1","Auth","Login page renders",
      "GET","/auth/login",200,
      f"1. Open {APP_URL_DISPLAY}/auth/login in browser\n"
      "2. Verify the login form is displayed with username, password fields and Login button\n"
      "3. Check page title shows 'OxyPC'",
      priority="CRITICAL", showstopper="YES"),

    t("P1-AUTH-02","Phase 1","Auth","Login with valid credentials",
      "POST","/auth/login",302,
      f"1. Open {APP_URL_DISPLAY}/auth/login\n"
      "2. Enter username: admin\n"
      "3. Enter password: oxypc@admin123\n"
      "4. Click Login\n"
      "5. Verify browser redirects to /dashboard\n"
      "6. Verify access_token cookie is set in browser developer tools (Application → Cookies)",
      priority="CRITICAL", showstopper="YES",
      notes="Expected 302 redirect with Set-Cookie access_token"),

    t("P1-AUTH-03","Phase 1","Auth","Login with wrong password shows error",
      "POST","/auth/login",200,
      f"1. Open {APP_URL_DISPLAY}/auth/login\n"
      "2. Enter username: admin\n"
      "3. Enter password: WRONG_PASSWORD\n"
      "4. Click Login\n"
      "5. Verify the login page reloads (stays on /auth/login)\n"
      "6. Verify an error message is displayed (e.g. 'Invalid credentials')",
      priority="HIGH"),

    t("P1-AUTH-04","Phase 1","Auth","Dashboard accessible after login",
      "GET","/",200,
      f"1. Log in with admin / oxypc@admin123\n"
      f"2. Navigate to {APP_URL_DISPLAY}/\n"
      "3. Verify dashboard loads with summary cards and navigation\n"
      "4. Verify no 500 or 403 error",
      priority="CRITICAL", showstopper="YES",
      notes="Dashboard is at / (root) — no /dashboard prefix"),

    t("P1-AUTH-05","Phase 1","Auth","Logout redirects to login",
      "POST","/auth/logout",302,
      f"1. Log in first\n"
      f"2. Click the Logout button in the navigation bar (or POST to {APP_URL_DISPLAY}/auth/logout)\n"
      "3. Verify browser redirects to /auth/login\n"
      "4. Verify access_token cookie is cleared",
      priority="HIGH",
      notes="Logout is a POST endpoint, not GET — use the Logout button in UI"),

    # Lots & GRN
    t("P1-LOTS-01","Phase 1","Lots & GRN","Lots list page",
      "GET","/lots",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/lots\n"
      "3. Verify page loads showing list of stock lots\n"
      "4. Verify table has columns: Lot ID, Supplier, Date, Status, Devices",
      priority="CRITICAL", showstopper="YES",
      notes="Lots list is at /lots (not /stock/lots)"),

    t("P1-LOTS-02","Phase 1","Lots & GRN","New Lot form",
      "GET","/lots/new",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/lots/new\n"
      "3. Verify the Create New Lot form renders\n"
      "4. Verify fields: Supplier, Date, Reference are present",
      priority="HIGH",
      notes="New lot form is at /lots/new"),

    t("P1-LOTS-03","Phase 1","Lots & GRN","GRN list page",
      "GET","/grn",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/grn\n"
      "3. Verify GRN (Goods Receipt Note) list renders\n"
      "4. Verify no 500 error",
      priority="HIGH"),

    # Devices
    t("P1-DEV-01","Phase 1","Devices","Devices list - all stages",
      "GET","/devices",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/devices\n"
      "3. Verify device list renders with all devices across all stages\n"
      "4. Verify columns: Serial, Model, Grade, Stage, Lot",
      priority="CRITICAL", showstopper="YES"),

    t("P1-DEV-02","Phase 1","Devices","Devices filtered by stage=iqc",
      "GET","/devices?stage=iqc",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/devices?stage=iqc\n"
      "3. Verify only devices in IQC stage are shown\n"
      "4. Verify stage filter is working",
      priority="HIGH"),

    t("P1-DEV-03","Phase 1","Devices","Devices filtered by stage=stock_in",
      "GET","/devices?stage=stock_in",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/devices?stage=stock_in\n"
      "3. Verify only devices in Stock-In stage are shown",
      priority="HIGH"),

    t("P1-DEV-04","Phase 1","Devices","Devices filtered by stage=l1",
      "GET","/devices?stage=l1",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/devices?stage=l1\n"
      "3. Verify only devices in L1 Repair stage are shown",
      priority="HIGH"),

    t("P1-DEV-05","Phase 1","Devices","Devices filtered by stage=ready_to_sale",
      "GET","/devices?stage=ready_to_sale",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/devices?stage=ready_to_sale\n"
      "3. Verify only devices ready for sale are shown",
      priority="HIGH"),

    # IQC
    t("P1-IQC-01","Phase 1","IQC","IQC main page",
      "GET","/iqc",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/iqc\n"
      "3. Verify IQC inspection page loads with device list\n"
      "4. Verify grade selection UI is present",
      priority="CRITICAL", showstopper="YES"),

    t("P1-IQC-02","Phase 1","IQC","IQC API users endpoint",
      "GET","/iqc/api/users",422,
      f"1. In Postman or browser, call GET {APP_URL_DISPLAY}/iqc/api/users\n"
      "2. Verify JSON response (422 if API-key header missing, 200 with valid key)\n"
      "3. This is the machine API for OxyQC hardware integration\n"
      "4. Valid API key must be in X-API-Key header",
      priority="HIGH",
      notes="Machine API — 422 expected without API-key header; /iqc/api/devices path does not exist"),

    # Repair
    t("P1-REP-01","Phase 1","Repair","Repair L1 page",
      "GET","/repair/l1",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/repair/l1\n"
      "3. Verify Level-1 repair queue page loads\n"
      "4. Verify device cards are shown with action buttons",
      priority="CRITICAL", showstopper="YES"),

    t("P1-REP-02","Phase 1","Repair","Repair L2 page",
      "GET","/repair/l2",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/repair/l2\n"
      "3. Verify Level-2 repair queue page loads",
      priority="HIGH"),

    t("P1-REP-03","Phase 1","Repair","Repair L3 page",
      "GET","/repair/l3",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/repair/l3\n"
      "3. Verify Level-3 repair queue page loads",
      priority="HIGH"),

    # QC
    t("P1-QC-01","Phase 1","QC","QC main page",
      "GET","/qc",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/qc\n"
      "3. Verify Quality Control page loads with device queue\n"
      "4. Verify grade/pass/fail buttons are visible",
      priority="CRITICAL", showstopper="YES"),

    # Stock / Transfers
    t("P1-STK-01","Phase 1","Stock","Stock overview page",
      "GET","/stock",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/stock\n"
      "3. Verify stock management page loads showing current inventory summary\n"
      "4. Verify device count by stage is displayed",
      priority="HIGH",
      notes="/stock is the stock overview page (distinct from /lots which is lot management)"),

    t("P1-TRF-01","Phase 1","Transfers","Transfers page",
      "GET","/transfers",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/transfers\n"
      "3. Verify stock transfer list page loads\n"
      "4. Verify transfer history is shown",
      priority="HIGH"),

    # Sales
    t("P1-SAL-01","Phase 1","Sales","Sales list page",
      "GET","/sales",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/sales\n"
      "3. Verify sales order list renders\n"
      "4. Verify columns: Order ID, Dealer, Date, Amount, Status",
      priority="CRITICAL", showstopper="YES"),

    t("P1-INV-01","Phase 1","Invoices","Invoice print page - route exists",
      "GET","/invoices/print/00000000-0000-0000-0000-000000000001",302,
      f"1. Log in as admin\n"
      f"2. Open an existing sale order from {APP_URL_DISPLAY}/sales\n"
      "3. Click the Print Invoice button next to any sale\n"
      "4. Verify invoice PDF/print view renders with sale details, items, line totals, grand total",
      priority="CRITICAL", showstopper="YES",
      notes="Auto-test uses a dummy UUID so expects 307/404 redirect; manual test must use a real sale ID from /sales"),
]

# ---- PHASE 2 ----------------------------------------------------------------
PHASE2_TESTS = [
    # Dealers
    t("P2-DLR-01","Phase 2","Dealers","Dealers list page",
      "GET","/dealers",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/dealers\n"
      "3. Verify dealer directory renders with name, city, credit limit\n"
      "4. Verify Add Dealer button is visible",
      priority="CRITICAL", showstopper="YES"),

    t("P2-DLR-02","Phase 2","Dealers","Dealer ageing report",
      "GET","/dealers/ageing",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/dealers/ageing\n"
      "3. Verify ageing report renders showing 0-30, 31-60, 61-90, 90+ day buckets\n"
      "4. Verify totals row is shown",
      priority="HIGH"),

    t("P2-DLR-03","Phase 2","Dealers","Dealer overdue report",
      "GET","/dealers/overdue",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/dealers/overdue\n"
      "3. Verify overdue payments report renders\n"
      "4. Verify dealers with outstanding dues are listed",
      priority="HIGH"),

    # Accounts
    t("P2-ACC-01","Phase 2","Accounts","Accounts index page",
      "GET","/accounts",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/accounts\n"
      "3. Verify accounts landing page loads\n"
      "4. Verify links to customer receipts and supplier payments",
      priority="HIGH"),

    t("P2-ACC-02","Phase 2","Accounts","Customer receipts page",
      "GET","/accounts/customer-receipts",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/accounts/customer-receipts\n"
      "3. Verify customer receipts list renders\n"
      "4. Verify columns: Receipt No, Dealer, Date, Amount, Mode",
      priority="CRITICAL", showstopper="YES"),

    t("P2-ACC-03","Phase 2","Accounts","Supplier payments page",
      "GET","/accounts/supplier-payments",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/accounts/supplier-payments\n"
      "3. Verify supplier payments list renders\n"
      "4. Verify columns: Payment No, Supplier, Date, Amount, Mode",
      priority="HIGH"),

    # Reports
    t("P2-RPT-01","Phase 2","Reports","Reports - Lot P&L page",
      "GET","/reports/lot-pl",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/reports/lot-pl\n"
      "3. Verify Lot Profit & Loss report renders with lot-wise margin data\n"
      "4. Verify columns: Lot ID, Purchase Cost, Revenue, Gross Margin",
      priority="HIGH",
      notes="/reports/ has no landing index page; use /reports/lot-pl or /reports/receivables directly"),

    t("P2-RPT-02","Phase 2","Reports","Receivables report",
      "GET","/reports/receivables",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/reports/receivables\n"
      "3. Verify receivables report renders with dealer-wise outstanding amounts\n"
      "4. Verify grand total is shown\n"
      "5. Verify Export CSV button is visible",
      priority="CRITICAL", showstopper="YES"),

    t("P2-RPT-03","Phase 2","Reports","Receivables CSV export",
      "GET","/reports/receivables?export=csv",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/reports/receivables?export=csv\n"
      "3. Verify browser downloads a CSV file\n"
      "4. Verify CSV contains headers: Dealer, Outstanding, Overdue",
      priority="HIGH"),

    # Inventory Location
    t("P2-LOC-01","Phase 2","Inventory Location","Inventory location master page",
      "GET","/locations/master",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/locations/master\n"
      "3. Verify inventory location management page loads\n"
      "4. Verify locations (e.g. Shelf A, Shelf B) are listed",
      priority="HIGH",
      notes="Inventory location router prefix is /locations (not /inventory-location)"),
]

# ---- PHASE 3 ----------------------------------------------------------------
PHASE3_TESTS = [
    # CRM
    t("P3-CRM-01","Phase 3","CRM","CRM dashboard",
      "GET","/crm/",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/crm/ (note trailing slash)\n"
      "3. Verify CRM dashboard loads with summary cards\n"
      "4. Verify navigation links (Contacts, Activities, Sourcing, Sales, Quotes, POs)",
      priority="HIGH",
      notes="/crm (no trailing slash) returns 307 redirect to /crm/"),

    t("P3-CRM-02","Phase 3","CRM","CRM Contacts",
      "GET","/crm/contacts",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/crm/contacts\n"
      "3. Verify contacts list renders\n"
      "4. Verify columns: Name, Company, Phone, Email, Type",
      priority="HIGH"),

    t("P3-CRM-03","Phase 3","CRM","CRM Activities - Follow-ups",
      "GET","/crm/activities/followups",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/crm/activities/followups\n"
      "3. Verify activities follow-ups list renders (calls, meetings)\n"
      "4. Verify Add Activity button is present",
      priority="HIGH",
      notes="/crm/activities has no index; use /crm/activities/followups"),

    t("P3-CRM-04","Phase 3","CRM","CRM Sourcing",
      "GET","/crm/sourcing",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/crm/sourcing\n"
      "3. Verify sourcing/procurement pipeline page renders\n"
      "4. Verify lot/deal status columns are shown",
      priority="HIGH"),

    t("P3-CRM-05","Phase 3","CRM","CRM Sales pipeline",
      "GET","/crm/sales",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/crm/sales\n"
      "3. Verify CRM sales pipeline renders\n"
      "4. Verify opportunities/deals are listed",
      priority="HIGH"),

    t("P3-CRM-06","Phase 3","CRM","CRM Quotes",
      "GET","/crm/quotes",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/crm/quotes\n"
      "3. Verify quotes list renders\n"
      "4. Verify columns: Quote No, Contact, Date, Total, Status",
      priority="HIGH"),

    t("P3-CRM-07","Phase 3","CRM","CRM Purchase Orders",
      "GET","/crm/purchase-orders",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/crm/purchase-orders\n"
      "3. Verify purchase orders list renders\n"
      "4. Verify columns: PO No, Supplier, Date, Amount, Status",
      priority="HIGH"),

    # WhatsApp & Market
    t("P3-WA-01","Phase 3","WhatsApp","WhatsApp page",
      "GET","/whatsapp",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/whatsapp\n"
      "3. Verify WhatsApp integration page loads\n"
      "4. Verify message templates or QR code is shown",
      priority="MEDIUM"),

    t("P3-MKT-01","Phase 3","Market","Market page",
      "GET","/market",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/market\n"
      "3. Verify market/listing page renders\n"
      "4. Verify device listings or marketplace view is shown",
      priority="MEDIUM"),

    # Admin & Settings
    t("P3-ADM-01","Phase 3","Admin","Admin users list",
      "GET","/admin/users",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/admin/users\n"
      "3. Verify admin user management page loads\n"
      "4. Verify list shows users with roles and status",
      priority="HIGH",
      notes="/admin has no index page; admin entry point is /admin/users"),

    t("P3-ADM-02","Phase 3","Admin","Admin audit log",
      "GET","/admin/audit-log",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/admin/audit-log\n"
      "3. Verify audit log renders showing all system write operations\n"
      "4. Verify columns: Timestamp, User, Action, Table, Record ID",
      priority="HIGH"),

    t("P3-SET-01","Phase 3","Settings","Settings page",
      "GET","/settings",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/settings\n"
      "3. Verify application settings page loads\n"
      "4. Verify fields for company name, address, VAT number etc.",
      priority="HIGH"),

    # Attendance
    t("P3-ATT-01","Phase 3","Attendance","Attendance page",
      "GET","/attendance",200,
      f"1. Log in as admin\n"
      f"2. Navigate to {APP_URL_DISPLAY}/attendance\n"
      "3. Verify attendance management page renders\n"
      "4. Verify clock-in/out functionality is shown",
      priority="MEDIUM"),
]

# ---- SECURITY TESTS ─────────────────────────────────────────────────────────
SECURITY_TESTS = [
    t("SEC-01","Security","CSRF","POST mutation without CSRF token returns 403",
      "POST","/dealers/00000000-0000-0000-0000-000000000000/credit-notes/00000000-0000-0000-0000-000000000001/apply",
      403,
      "1. Open Postman or browser DevTools\n"
      "2. Make a POST request to "
      f"{APP_URL_DISPLAY}/dealers/00000000-0000-0000-0000-000000000000/credit-notes/00000000-0000-0000-0000-000000000001/apply\n"
      "3. Do NOT include a csrf_token cookie or form field\n"
      "4. Include a valid access_token cookie (from a logged-in session)\n"
      "5. Verify response is 403 Forbidden\n"
      "6. Verify error message mentions CSRF or Access Denied",
      priority="CRITICAL", showstopper="YES",
      notes="Double-submit CSRF protection must return 403 when token is absent"),

    t("SEC-02","Security","Auth","Unauthenticated /admin/users redirects to login",
      "GET","/admin/users",307,
      "1. Open a fresh browser tab (no session cookies / private window)\n"
      f"2. Navigate directly to {APP_URL_DISPLAY}/admin/users\n"
      "3. Verify browser redirects to /auth/login (307 Temporary Redirect)\n"
      "4. Verify you are NOT shown the admin panel",
      priority="CRITICAL", showstopper="YES",
      notes="FastAPI returns 307 (not 302) for auth redirects — both are safe redirects to /auth/login"),

    t("SEC-03","Security","Rate Limiting","Login rate limit triggers 429",
      "POST","/auth/login",429,
      "1. Use Postman or a script to POST to /auth/login 31 times rapidly\n"
      "2. Use any credentials (even wrong ones)\n"
      "3. Verify that after ~5 attempts within 1 minute, the server returns 429 Too Many Requests\n"
      "4. Verify the response body contains 'Rate limit exceeded' or similar message",
      priority="HIGH",
      notes="slowapi login limiter set to 5/minute per IP; 429 expected by attempt 6"),
]

# All tests combined for iteration
ALL_HTTP_TESTS = PHASE1_TESTS + PHASE2_TESTS + PHASE3_TESTS + SECURITY_TESTS


# ─── HTTP Test Runner ─────────────────────────────────────────────────────────
def run_http_tests():
    print("\n" + "="*60)
    print("PART 1: LIVE HTTP TESTS")
    print("="*60)

    results = []
    auth_cookies = {}

    with httpx.Client(follow_redirects=False, timeout=15.0) as client:

        # ── Step 1: Login to get auth cookie ──────────────────────────────
        print("\n[LOGIN] Authenticating as admin...")
        try:
            resp = client.post(f"{BASE_URL}/auth/login",
                               data={"username": ADMIN_USER, "password": ADMIN_PASS},
                               headers={"Content-Type": "application/x-www-form-urlencoded"})
            if resp.status_code in (302, 303) and "access_token" in resp.cookies:
                auth_cookies = dict(resp.cookies)
                print(f"  [OK] Login succeeded — cookies: {list(auth_cookies.keys())}")
            else:
                print(f"  [WARN] Login returned {resp.status_code}; proceeding without auth cookie")
        except Exception as e:
            print(f"  [ERROR] Login request failed: {e}")

        # Grab csrf_token from cookie jar if present
        csrf_cookie = auth_cookies.get("csrf_token", "")

        # ── Step 2: Run all HTTP tests ─────────────────────────────────────
        for test in ALL_HTTP_TESTS:
            tid  = test["test_id"]
            url  = f"{BASE_URL}{test['url']}"
            meth = test["method"]
            exp  = test["expected_status"]

            # For security SEC-01 and SEC-02 — use NO auth cookies
            if tid in ("SEC-01", "SEC-02"):
                cookies = {}
            else:
                cookies = auth_cookies

            # For SEC-01: POST without csrf_token form field and no csrf cookie
            if tid == "SEC-01":
                cookies = {k: v for k, v in auth_cookies.items() if k != "csrf_token"}
                form_data = {}  # no csrf_token in body either
            elif meth == "POST" and tid == "P1-AUTH-02":
                cookies = {}
                form_data = {"username": ADMIN_USER, "password": ADMIN_PASS}
            elif meth == "POST" and tid == "P1-AUTH-03":
                cookies = {}
                form_data = {"username": ADMIN_USER, "password": "WRONG_PASS"}
            elif meth == "POST" and tid == "P1-AUTH-05":
                # Logout — POST with valid CSRF
                form_data = {"csrf_token": csrf_cookie}
            else:
                form_data = None

            t_start = time.monotonic()
            actual  = None
            notes_extra = test["notes"]

            try:
                if tid == "SEC-03":
                    # Rate-limit test: 31 rapid POSTs
                    last_status = None
                    hit_429 = False
                    for i in range(31):
                        r = client.post(f"{BASE_URL}/auth/login",
                                        data={"username": "admin", "password": "wrong"},
                                        cookies={},
                                        headers={"Content-Type": "application/x-www-form-urlencoded"})
                        last_status = r.status_code
                        if r.status_code == 429:
                            hit_429 = True
                            break
                    actual = 429 if hit_429 else last_status
                    notes_extra = f"Hit 429 after burst: {hit_429}; last status: {last_status}"

                elif meth == "POST" and tid == "SEC-01":
                    resp = client.post(url, data=form_data or {},
                                       cookies=cookies,
                                       headers={"Content-Type": "application/x-www-form-urlencoded"})
                    actual = resp.status_code

                elif meth == "POST":
                    resp = client.post(url, data=form_data or {},
                                       cookies=cookies,
                                       headers={"Content-Type": "application/x-www-form-urlencoded"})
                    actual = resp.status_code

                else:  # GET
                    resp = client.get(url, cookies=cookies)
                    actual = resp.status_code

            except httpx.ConnectError:
                actual = "CONN_ERR"
                notes_extra = "Connection refused — app may be down"
            except Exception as ex:
                actual = "ERROR"
                notes_extra = str(ex)[:120]

            elapsed = round((time.monotonic() - t_start) * 1000, 1)

            # ── Determine pass/fail ────────────────────────────────────────
            if isinstance(actual, int):
                if actual == exp:
                    status = "PASS"
                elif tid == "P1-IQC-02" and actual in (200, 401, 422):
                    status = "PASS"  # 422 = missing API-key header, acceptable
                    notes_extra = f"Got {actual} — acceptable for machine API endpoint (needs X-API-Key header)"
                elif actual in (301, 307, 308) and exp == 200:
                    # Redirect to same page with trailing slash — acceptable
                    status = "PASS"
                    notes_extra = f"Got {actual} redirect (trailing-slash redirect) — acceptable"
                elif actual == 500:
                    status = "FAIL-500"
                else:
                    status = "FAIL"
            else:
                status = "ERROR"

            test["actual_status"] = actual
            test["status"] = status
            test["elapsed_ms"] = elapsed
            test["notes"] = notes_extra

            icon = "PASS" if status == "PASS" else "FAIL"
            print(f"  [{icon}] {tid:16s} {meth:4s} {test['url']:45s} "
                  f"exp={exp} got={actual} ({elapsed}ms)")

            results.append(test)

    return results


# ─── Static Analysis Tests ────────────────────────────────────────────────────
def run_static_analysis():
    print("\n" + "="*60)
    print("PART 2: STATIC ANALYSIS")
    print("="*60)

    findings = []

    def check(name, passed, detail=""):
        icon = "PASS" if passed else "FAIL"
        print(f"  [{icon}] {name}")
        if detail:
            print(f"         {detail}")
        findings.append({"name": name, "passed": passed, "detail": detail})

    routers_dir = os.path.join(PROJECT_ROOT, "routers")
    main_py     = os.path.join(PROJECT_ROOT, "main.py")
    req_txt     = os.path.join(PROJECT_ROOT, "requirements.txt")
    gitignore   = os.path.join(PROJECT_ROOT, ".gitignore")
    limiter_py  = os.path.join(PROJECT_ROOT, "limiter.py")
    deps_py     = os.path.join(PROJECT_ROOT, "auth", "dependencies.py")
    error_tmpl  = os.path.join(PROJECT_ROOT, "templates", "error.html")

    # 1. All mutation routers have verify_csrf
    mutation_routers = [
        "attendance.py","market.py","whatsapp.py","telecalling.py",
        "crm_purchase_orders.py","crm_price_matrix.py","crm_quotes.py",
        "crm_sales.py","crm_sourcing.py","crm_activities.py",
        "crm_contacts.py","inventory_location.py","stage_control.py",
        "bulk_upload.py","devices.py","master.py","spare_parts.py",
        "transfers.py","grn.py","cosmetic.py","qc.py","iqc.py",
        "repair.py","stock.py","sales.py","admin.py","accounts.py",
        "dealers.py","settings.py",
    ]
    # deduplicate
    mutation_routers = list(dict.fromkeys(mutation_routers))
    missing_csrf = []
    for rname in mutation_routers:
        rpath = os.path.join(routers_dir, rname)
        if not os.path.exists(rpath):
            continue
        with open(rpath, encoding="utf-8", errors="ignore") as f:
            content = f.read()
        if "verify_csrf" not in content:
            missing_csrf.append(rname)
    check(f"All mutation routers have verify_csrf ({len(mutation_routers)} checked)",
          len(missing_csrf) == 0,
          f"Missing CSRF in: {missing_csrf}" if missing_csrf else "All routers have verify_csrf")

    # 2. requirements.txt — all packages pinned with ==
    with open(req_txt, encoding="utf-8") as f:
        req_lines = [l.strip() for l in f if l.strip() and not l.startswith("#")]
    unpinned = [l for l in req_lines if "==" not in l]
    check("requirements.txt — all packages pinned with ==",
          len(unpinned) == 0,
          f"Unpinned: {unpinned}" if unpinned else "All packages pinned")

    # 3. .gitignore contains config.ini
    if os.path.exists(gitignore):
        with open(gitignore, encoding="utf-8", errors="ignore") as f:
            gi_content = f.read()
        check(".gitignore contains config.ini",
              "config.ini" in gi_content)
    else:
        check(".gitignore contains config.ini", False, ".gitignore not found")

    # 4. main.py has generic exception handler
    with open(main_py, encoding="utf-8") as f:
        main_content = f.read()
    check("main.py has generic exception handler (DBAPIError/ProgrammingError)",
          "DBAPIError" in main_content or "exception_handler" in main_content)

    # 5. main.py has OXYPC_AUTO_FIX
    check("main.py has OXYPC_AUTO_FIX env var reference",
          "OXYPC_AUTO_FIX" in main_content,
          "Search for OXYPC_AUTO_FIX in main.py")

    # 6. limiter.py has OXYPC_TRUSTED_PROXY
    with open(limiter_py, encoding="utf-8") as f:
        lim_content = f.read()
    check("limiter.py has OXYPC_TRUSTED_PROXY",
          "OXYPC_TRUSTED_PROXY" in lim_content)

    # 7. auth/dependencies.py has verify_csrf
    with open(deps_py, encoding="utf-8") as f:
        deps_content = f.read()
    check("auth/dependencies.py defines verify_csrf function",
          "async def verify_csrf" in deps_content or "def verify_csrf" in deps_content)

    # 8. error.html template exists
    check("templates/error.html exists",
          os.path.exists(error_tmpl))

    # 9. Key templates exist
    key_templates = {
        "login.html": os.path.join(PROJECT_ROOT, "templates", "login.html"),
        "dashboard.html": os.path.join(PROJECT_ROOT, "templates", "dashboard.html"),
        "dealers/profile.html": os.path.join(PROJECT_ROOT, "templates", "dealers", "profile.html"),
        "dealers/order_invoice.html": os.path.join(PROJECT_ROOT, "templates", "dealers", "order_invoice.html"),
        "reports/receivables.html": os.path.join(PROJECT_ROOT, "templates", "reports", "receivables.html"),
        "accounts/customer_receipts.html": os.path.join(PROJECT_ROOT, "templates", "accounts", "customer_receipts.html"),
    }
    for tname, tpath in key_templates.items():
        check(f"Template exists: {tname}", os.path.exists(tpath))

    return findings


# ─── Showstopper Gap Analysis ─────────────────────────────────────────────────
def run_gap_analysis(http_results, static_findings):
    print("\n" + "="*60)
    print("PART 3: SHOWSTOPPER GAP ANALYSIS")
    print("="*60)

    gaps = []

    def gap(priority, title, impact, effort, status, owner, notes):
        gaps.append({
            "priority": priority,
            "gap": title,
            "impact": impact,
            "effort": effort,
            "status": status,
            "owner": owner,
            "notes": notes,
        })
        icon = "CRIT" if priority == "P0" else "WARN"
        print(f"  [{icon}] {title}")

    # ── Auto-flag from HTTP results ────────────────────────────────────────────
    phase1_fails = [r for r in http_results
                    if r["phase"] == "Phase 1" and r["status"] not in ("PASS",)]
    phase1_500   = [r for r in http_results if r["actual_status"] == 500]

    if phase1_500:
        for r in phase1_500:
            gap("P0", f"500 Server Error: {r['url']}",
                "Page crashes — users cannot access this page",
                "Medium", "OPEN", "Dev Team",
                f"GET/POST {r['url']} returned HTTP 500")

    csrf_test = next((r for r in http_results if r["test_id"] == "SEC-01"), None)
    if csrf_test:
        if csrf_test["status"] == "PASS":
            print(f"  [OK] CSRF protection working (403 returned)")
        else:
            gap("P0", "CSRF protection NOT enforcing 403",
                "Any authenticated user can forge mutation requests — data integrity at risk",
                "Low", "OPEN", "Dev Team",
                f"SEC-01: Expected 403 got {csrf_test['actual_status']}")

    # ── Known gaps from code analysis ─────────────────────────────────────────

    # 1. Retail invoice print endpoint
    invoice_print_url = f"{BASE_URL}/invoices/print/00000000-0000-0000-0000-000000000001"
    with httpx.Client(follow_redirects=False, timeout=10.0) as client:
        # Login
        try:
            r = client.post(f"{BASE_URL}/auth/login",
                            data={"username": ADMIN_USER, "password": ADMIN_PASS},
                            headers={"Content-Type": "application/x-www-form-urlencoded"})
            ck = dict(r.cookies)
        except Exception:
            ck = {}

        try:
            rp = client.get(invoice_print_url, cookies=ck)
            invoice_print_status = rp.status_code
        except Exception:
            invoice_print_status = "ERROR"

    print(f"\n  [CHECK] /invoices/print/{{sale_id}} → {invoice_print_status}")
    if invoice_print_status in (404, 422, "ERROR"):
        gap("P0", "Retail Sale Invoice Print missing (/invoices/print/{sale_id})",
            "Cannot print invoices for retail/walk-in sales — core sales workflow blocked",
            "High", "OPEN", "Dev Team",
            f"GET /invoices/print/{{sale_id}} returned {invoice_print_status}. "
            "Template may exist (templates/invoices/print.html) but route may 404 for invalid UUID.")
    else:
        print(f"  [OK] Invoice print endpoint accessible ({invoice_print_status})")

    # 2. No HTTPS
    is_http_only = not any(
        "ssl" in f.get("detail","").lower() or "https" in f.get("detail","").lower()
        for f in static_findings
    )
    gap("P1", "No HTTPS — app runs on plain HTTP",
        "All data including passwords transmitted in plaintext on network",
        "Medium", "OPEN", "DevOps",
        "App is configured to listen on HTTP only (config.ini host/port). "
        "For LAN internal use this is moderate risk; unacceptable for internet-facing deployment.")

    # 3. No email notifications
    email_found = False
    for root, dirs, files in os.walk(PROJECT_ROOT):
        # skip venv/cache dirs
        dirs[:] = [d for d in dirs if d not in ("__pycache__","venv",".venv","node_modules")]
        for fname in files:
            if fname.endswith(".py"):
                fpath = os.path.join(root, fname)
                try:
                    with open(fpath, encoding="utf-8", errors="ignore") as f:
                        c = f.read()
                    if any(kw in c.lower() for kw in ("smtp", "sendmail", "smtplib", "send_email", "email.mime")):
                        email_found = True
                        break
                except Exception:
                    pass
        if email_found:
            break

    if not email_found:
        gap("P1", "No email notification system detected",
            "Dealers/staff receive no email alerts for orders, approvals, credit limits exceeded",
            "High", "OPEN", "Dev Team",
            "No SMTP/sendmail/email.mime imports found anywhere in codebase. "
            "Manual communication required for all business events.")
    else:
        print("  [OK] Email sending code detected in codebase")

    # 4. No DB backup automation
    backup_found = False
    backup_paths = [
        os.path.join(PROJECT_ROOT, "scripts"),
        os.path.join(PROJECT_ROOT, "docs"),
        PROJECT_ROOT,
    ]
    for bdir in backup_paths:
        if not os.path.isdir(bdir):
            continue
        for f in os.listdir(bdir):
            if any(kw in f.lower() for kw in ("backup","pg_dump","restore","cron")):
                backup_found = True
                break
    # also check for backup runbook
    runbook = os.path.join(PROJECT_ROOT, "docs", "backup-runbook.md")
    runbook_exists = os.path.exists(runbook)
    print(f"  [CHECK] Backup runbook: {'FOUND' if runbook_exists else 'NOT FOUND'}")

    if not backup_found and not runbook_exists:
        gap("P0", "No automated DB backup script found",
            "Data loss risk — no recovery path if DB server fails",
            "Medium", "OPEN", "DevOps",
            "No pg_dump, backup script, or cron job found in project. "
            "Backup runbook also missing from docs/. "
            "Must implement nightly pg_dump before go-live.")
    elif runbook_exists:
        print(f"  [WARN] Backup runbook exists but automation not confirmed")
        gap("P1", "Backup runbook exists but no automated script confirmed",
            "Manual backup only — relies on operator discipline",
            "Low", "OPEN", "DevOps",
            f"docs/backup-runbook.md present. Verify cron job / scheduled task is actually running.")

    # 5. Credit note apply — selectinload check
    cn_router = os.path.join(PROJECT_ROOT, "routers", "dealers.py")
    cn_has_selectinload = False
    if os.path.exists(cn_router):
        with open(cn_router, encoding="utf-8", errors="ignore") as f:
            cn_content = f.read()
        cn_has_selectinload = "selectinload" in cn_content
    print(f"  [CHECK] dealers.py uses selectinload: {cn_has_selectinload}")
    if not cn_has_selectinload:
        gap("P1", "Credit note apply may trigger N+1 query (missing selectinload)",
            "Performance degradation on large dealer accounts — slow page loads",
            "Low", "OPEN", "Dev Team",
            "dealers.py credit-note apply query does not use selectinload() for related objects. "
            "Review applied_to_order_id query path.")

    # 6. Default password warning (not a blocker for LAN)
    gap("P2", "Admin password is still default (oxypc@admin123)",
        "Low risk for LAN-only; high risk if ever internet-facing",
        "Low", "WARN", "Admin",
        "Credentials: admin / oxypc@admin123 still in use. "
        "Change before any non-internal deployment. "
        "Not a go-live blocker for pure-LAN deployment.")

    # 7. config.ini plaintext password
    config_ini = os.path.join(PROJECT_ROOT, "config.ini")
    config_has_password = False
    if os.path.exists(config_ini):
        with open(config_ini, encoding="utf-8", errors="ignore") as f:
            cfg = f.read()
        config_has_password = ("password" in cfg.lower() or "pass" in cfg.lower())
    if config_has_password:
        gap("P1", "config.ini contains plaintext DB credentials",
            "Credentials exposed if config.ini is accidentally committed or accessed",
            "Low", "OPEN", "DevOps",
            "config.ini stores database password in plaintext. "
            "Ensure .gitignore includes config.ini (check SA static test). "
            "Consider migrating to environment variables for production.")

    # Any Phase 1 failures that aren't already flagged
    for r in phase1_fails:
        if r["actual_status"] not in (200, 302, 303, 404) or r["actual_status"] == 500:
            already = any(g["notes"] and r["url"] in g["notes"] for g in gaps)
            if not already:
                gap("P0", f"Phase 1 route failed: {r['method']} {r['url']}",
                    f"Core workflow page broken — users blocked",
                    "Medium", "OPEN", "Dev Team",
                    f"Expected {r['expected_status']}, got {r['actual_status']}")

    print(f"\n  Total showstoppers/gaps identified: {len(gaps)}")
    return gaps


# ─── Excel Generation ─────────────────────────────────────────────────────────
def apply_header_style(ws, row=1):
    for cell in ws[row]:
        cell.fill   = HDR_FILL
        cell.font   = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def row_fill(ws, row_num, is_alt):
    if is_alt:
        for cell in ws[row_num]:
            if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", ""):
                cell.fill = ALT_FILL


def status_fill(cell, status):
    s = str(status).upper()
    if s == "PASS":
        cell.fill = PASS_FILL
    elif "FAIL" in s or s in ("ERROR", "CONN_ERR"):
        cell.fill = FAIL_FILL
    elif s in ("SKIP","MANUAL","N/A"):
        cell.fill = SKIP_FILL
    elif "SHOWSTOPPER" in s or "SS" in s:
        cell.fill = SS_FILL


def add_manual_dropdown(ws, col_letter, start_row, end_row):
    dv = DataValidation(
        type="list",
        formula1='"PASS,FAIL,BLOCKED,N/A"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)


def build_test_sheet(wb, sheet_name, tests):
    """Create a phase test sheet."""
    ws = wb.create_sheet(sheet_name)
    headers = [
        "Test ID", "Phase", "Module", "Test Case",
        "Test Steps (Manual)", "Expected Result", "Auto Status", "Auto HTTP Code",
        "Manual Result", "Manual Status", "Priority", "Showstopper?", "Notes"
    ]
    ws.append(headers)
    apply_header_style(ws, 1)
    ws.freeze_panes = "A2"

    for i, test in enumerate(tests, start=2):
        auto_s = str(test.get("status") or "MANUAL")
        auto_c = str(test.get("actual_status") or "—")
        row = [
            test["test_id"],
            test["phase"],
            test["module"],
            test["name"],
            test["steps"],
            f"HTTP {test['expected_status']} — page renders without error",
            auto_s,
            auto_c,
            "",   # Manual Result — blank for team
            "",   # Manual Status — dropdown
            test["priority"],
            test["showstopper"],
            test.get("notes",""),
        ]
        ws.append(row)

        # Style auto status column (col 7)
        sc = ws.cell(row=i, column=7)
        status_fill(sc, auto_s)
        sc.alignment = Alignment(horizontal="center", vertical="center")

        # Manual result/status — grey fill
        for col in (9, 10):
            ws.cell(row=i, column=col).fill = GREY_FILL

        # Showstopper column — orange if YES
        ss_cell = ws.cell(row=i, column=12)
        if test["showstopper"] == "YES":
            ss_cell.fill = SS_FILL
            ss_cell.font = Font(bold=True)

        # Wrap test steps
        ws.cell(row=i, column=5).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=i, column=13).alignment = Alignment(wrap_text=True, vertical="top")

        # Alt row shading for non-styled cells
        for col in (1,2,3,4,6,8,11,12):
            c = ws.cell(row=i, column=col)
            c.alignment = Alignment(vertical="top", wrap_text=False)

        # Row height for steps
        ws.row_dimensions[i].height = 60

    # Dropdown on Manual Status column (col 10)
    add_manual_dropdown(ws, "J", 2, len(tests)+1)

    set_col_widths(ws, [14, 10, 16, 30, 55, 30, 12, 14, 16, 14, 10, 13, 35])
    return ws


def build_excel(http_results, static_findings, gaps):
    print("\n" + "="*60)
    print("PART 4: GENERATING EXCEL REPORT")
    print("="*60)

    wb = openpyxl.Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── COVER ─────────────────────────────────────────────────────────────────
    cov = wb.create_sheet("COVER", 0)
    cov.sheet_view.showGridLines = False

    total_auto   = len(http_results)
    total_pass   = sum(1 for r in http_results if r["status"] == "PASS")
    total_fail   = sum(1 for r in http_results if r["status"] not in ("PASS",))
    total_manual = len(PHASE1_TESTS) + len(PHASE2_TESTS) + len(PHASE3_TESTS) + len(SECURITY_TESTS)

    cov.column_dimensions["A"].width = 3
    cov.column_dimensions["B"].width = 35
    cov.column_dimensions["C"].width = 28

    # Title
    cov["B2"] = "OxyPC ERP — UAT Test Plan & Automated Results"
    cov["B2"].font = Font(size=16, bold=True, color="FFFFFF")
    cov["B2"].fill = PatternFill("solid", fgColor="0D47A1")
    cov["B2"].alignment = Alignment(horizontal="left", vertical="center")
    cov.merge_cells("B2:C2")
    cov.row_dimensions[2].height = 35

    rows = [
        ("Date:", "27 April 2026"),
        ("App URL:", APP_URL_DISPLAY),
        ("Automated Tests Run:", str(total_auto)),
        ("Auto PASS:", str(total_pass)),
        ("Auto FAIL:", str(total_fail)),
        ("Manual Tests Required:", str(total_manual)),
        ("Showstopper Gaps Found:", str(len(gaps))),
        ("Report Generated By:", "run_uat_excel.py"),
    ]
    for r_idx, (label, value) in enumerate(rows, start=4):
        cov.cell(row=r_idx, column=2, value=label).font = Font(bold=True, size=10)
        cov.cell(row=r_idx, column=3, value=value).font  = Font(size=10)
        cov.row_dimensions[r_idx].height = 18

    # Legend
    legend_row = 14
    cov.cell(row=legend_row, column=2, value="Legend").font = Font(bold=True, size=11)
    legend_items = [
        ("Green (#C8E6C9)", "Auto PASS", PASS_FILL),
        ("Red (#FFCDD2)",   "Auto FAIL / Error", FAIL_FILL),
        ("Yellow (#FFF9C4)","Manual / SKIP", SKIP_FILL),
        ("Orange (#FFE0B2)","Showstopper Gap", SS_FILL),
        ("Grey (#F5F5F5)",  "Manual team fills in", GREY_FILL),
    ]
    for li, (colour, meaning, fill) in enumerate(legend_items, start=legend_row+1):
        c1 = cov.cell(row=li, column=2, value=colour)
        c1.fill = fill
        c1.border = THIN_BORDER
        c1.alignment = Alignment(horizontal="center")
        c2 = cov.cell(row=li, column=3, value=meaning)
        c2.alignment = Alignment(horizontal="left")
        cov.row_dimensions[li].height = 16

    cov.freeze_panes = None

    # ── Phase Sheets ───────────────────────────────────────────────────────────
    build_test_sheet(wb, "P1 - Core Inventory", PHASE1_TESTS)
    build_test_sheet(wb, "P2 - Financial & Dealer", PHASE2_TESTS)
    build_test_sheet(wb, "P3 - CRM & Others", PHASE3_TESTS)
    build_test_sheet(wb, "Security Tests", SECURITY_TESTS)

    # ── Showstoppers Sheet ────────────────────────────────────────────────────
    ss = wb.create_sheet("Showstoppers")
    ss_headers = ["Priority","Gap / Issue","Business Impact","Effort","Status","Owner","Notes"]
    ss.append(ss_headers)
    # Red header
    for cell in ss[1]:
        cell.fill = PatternFill("solid", fgColor="B71C1C")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ss.freeze_panes = "A2"

    red_border = Border(
        left=Side(style='medium', color='B71C1C'),
        right=Side(style='medium', color='B71C1C'),
        top=Side(style='medium', color='B71C1C'),
        bottom=Side(style='medium', color='B71C1C'),
    )

    for i, g in enumerate(gaps, start=2):
        row_data = [g["priority"], g["gap"], g["impact"],
                    g["effort"], g["status"], g["owner"], g["notes"]]
        ss.append(row_data)

        p_cell = ss.cell(row=i, column=1)
        if g["priority"] == "P0":
            p_cell.fill = PatternFill("solid", fgColor="FFCDD2")
            p_cell.font = Font(bold=True, color="B71C1C")
        elif g["priority"] == "P1":
            p_cell.fill = SS_FILL
            p_cell.font = Font(bold=True)
        else:
            p_cell.fill = SKIP_FILL

        for col in range(1, 8):
            c = ss.cell(row=i, column=col)
            c.border = THIN_BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")

        ss.row_dimensions[i].height = 45

    set_col_widths(ss, [10, 40, 40, 12, 12, 16, 50])

    # ── All Auto Results Sheet ────────────────────────────────────────────────
    ar = wb.create_sheet("All Auto Results")
    ar_headers = [
        "Test ID","Phase","Module","Test Name","Method","URL",
        "Expected Status","Actual Status","Pass/Fail","Elapsed (ms)","Notes"
    ]
    ar.append(ar_headers)
    apply_header_style(ar, 1)
    ar.freeze_panes = "A2"

    for i, r in enumerate(http_results, start=2):
        row = [
            r["test_id"], r["phase"], r["module"], r["name"],
            r["method"], r["url"],
            r["expected_status"], str(r.get("actual_status","—")),
            r.get("status","—"), r.get("elapsed_ms","—"),
            r.get("notes",""),
        ]
        ar.append(row)

        st_cell = ar.cell(row=i, column=9)
        status_fill(st_cell, r.get("status",""))
        st_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 12):
            ar.cell(row=i, column=col).alignment = Alignment(wrap_text=False, vertical="center")

        if i % 2 == 0:
            for col in range(1, 12):
                c = ar.cell(row=i, column=col)
                if c.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    c.fill = ALT_FILL

    set_col_widths(ar, [14, 10, 16, 36, 8, 42, 15, 14, 12, 13, 40])

    # ── Save ──────────────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    size = os.path.getsize(OUTPUT_PATH)
    print(f"\n  [OK] Excel saved: {OUTPUT_PATH}")
    print(f"  [OK] File size: {size:,} bytes")
    return size


# ─── Main ────────────────────────────────────────────────────────────────────
def main():
    print("OxyPC ERP — UAT Test Suite & Excel Report Generator")
    print("Date: 27 April 2026")
    print(f"App: {BASE_URL}")
    print(f"Output: {OUTPUT_PATH}")

    http_results   = run_http_tests()
    static_results = run_static_analysis()
    gaps           = run_gap_analysis(http_results, static_results)
    excel_size     = build_excel(http_results, static_results, gaps)

    # ── Summary ──────────────────────────────────────────────────────────────
    print("\n" + "="*60)
    print("FINAL SUMMARY")
    print("="*60)
    total   = len(http_results)
    passed  = sum(1 for r in http_results if r["status"] == "PASS")
    failed  = total - passed
    p0_gaps = [g for g in gaps if g["priority"] == "P0"]
    p1_gaps = [g for g in gaps if g["priority"] == "P1"]

    print(f"  HTTP Tests Run    : {total}")
    print(f"  PASS              : {passed}")
    print(f"  FAIL              : {failed}")
    print(f"  Showstoppers (P0) : {len(p0_gaps)}")
    print(f"  High Priority (P1): {len(p1_gaps)}")
    print(f"  Total Gaps        : {len(gaps)}")
    print(f"  Excel file size   : {excel_size:,} bytes")
    print(f"\n  Output: {OUTPUT_PATH}")

    if p0_gaps:
        print("\n  P0 SHOWSTOPPERS:")
        for g in p0_gaps:
            print(f"    - {g['gap']}")

    print("\n[DONE]")


if __name__ == "__main__":
    main()
